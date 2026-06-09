import streamlit as st
import pandas as pd
import os
from datetime import datetime
import calendar

# ─────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────
DATA_DIR = "data"
ARCHIVO = os.path.join(DATA_DIR, "parametrosMesuales.xlsx")
HOJA = "Hoja2"

st.set_page_config(
    page_title="Rex+ | Parámetros Mensuales",
    page_icon="📋",
    layout="wide"
)

# ─────────────────────────────────────────────
# ESTILOS REX+
# ─────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

.rex-header {
    background: linear-gradient(135deg, #2d4a7a 0%, #1a2f50 100%);
    padding: 1.2rem 2rem;
    border-radius: 12px;
    margin-bottom: 1.5rem;
    display: flex;
    align-items: center;
    gap: 1rem;
}
.rex-logo {
    background: white;
    border-radius: 8px;
    padding: 4px 10px;
    font-weight: 700;
    font-size: 1.1rem;
    color: #1a2f50;
    letter-spacing: -0.5px;
}
.rex-logo span { color: #3b9fd1; }
.rex-title {
    color: white;
    font-size: 1.25rem;
    font-weight: 600;
}
.rex-badge {
    background: #3b9fd1;
    color: white;
    font-size: 0.65rem;
    font-weight: 700;
    padding: 3px 8px;
    border-radius: 20px;
    letter-spacing: 1px;
    margin-left: auto;
}
.section-card {
    background: #f8fafc;
    border: 1px solid #e2e8f0;
    border-radius: 10px;
    padding: 1.2rem 1.5rem;
    margin-bottom: 1rem;
}
.section-title {
    font-size: 0.85rem;
    font-weight: 700;
    color: #2d4a7a;
    text-transform: uppercase;
    letter-spacing: 0.8px;
    margin-bottom: 1rem;
    padding-bottom: 0.5rem;
    border-bottom: 2px solid #3b9fd1;
}
.stButton > button {
    background: linear-gradient(135deg, #2d4a7a, #1a2f50) !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    padding: 0.5rem 1.5rem !important;
}
.stButton > button:hover {
    background: linear-gradient(135deg, #3b9fd1, #2d4a7a) !important;
    transform: translateY(-1px);
}
.success-msg {
    background: #f0fdf4;
    border: 1px solid #86efac;
    border-radius: 8px;
    padding: 0.8rem 1rem;
    color: #166534;
    font-weight: 500;
}
.info-pill {
    display: inline-block;
    background: #eff6ff;
    color: #1d4ed8;
    font-size: 0.75rem;
    font-weight: 600;
    padding: 2px 10px;
    border-radius: 20px;
    margin-left: 8px;
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────────
st.markdown("""
<div class="rex-header">
    <div class="rex-logo">Rex<span>+</span></div>
    <div class="rex-title">Parámetros Mensuales</div>
    <div class="rex-badge">MANTENCIÓN</div>
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# ETIQUETAS LEGIBLES POR COLUMNA
# ─────────────────────────────────────────────
LABELS = {
    "mes_Proc":           "Mes de proceso (aaaa-mm)",
    "uf_Mes":             "UF del mes ($)",
    "topeImp_Uf_afp":     "Tope imponible AFP (UF)",
    "topeImp_pesos_afp":  "Tope imponible AFP ($)",
    "topeCes_Uf":         "Tope cesantía (UF)",
    "topeCes_pesos":      "Tope cesantía ($)",
    "sis":                "SIS (%)",
    "factor_sis":         "Factor SIS (decimal)",
    "topeSalud_Uf":       "Tope salud (UF)",
    "topeSalud_pesos":    "Tope salud ($)",
    "imm":                "IMM ($)",
    "topeGratif":         "Tope gratificación ($)",
    "monto_Utm":          "Monto UTM ($)",
    "ult_Diames":         "Último día del mes",
    "aporte_Ccaf":        "Aporte CCAF (%)",
    "aporte_Fonasa":      "Aporte FONASA (%)",
    "Formato Fecha":      "Fecha formato (dd/mm/aaaa)",
    "Aporte AFP":         "Aporte AFP (%)",
    "Seg Social Exp vida":"Seg. social / Exp. vida (%)",
}

# Columnas numéricas editables (no mes_Proc ni Formato Fecha)
COLS_NUM = [c for c in LABELS if c not in ("mes_Proc", "Formato Fecha")]

# ─────────────────────────────────────────────
# CARGAR DATOS
# ─────────────────────────────────────────────
@st.cache_data(ttl=0)
def cargar_datos():
    df = pd.read_excel(ARCHIVO, sheet_name=HOJA, dtype={"mes_Proc": str})
    df["mes_Proc"] = df["mes_Proc"].astype(str).str.strip()
    return df

def guardar_datos(df: pd.DataFrame):
    from openpyxl import load_workbook
    wb = load_workbook(ARCHIVO)
    ws = wb[HOJA]
    # Reescribir completamente la hoja con los datos del DataFrame
    ws.delete_rows(2, ws.max_row)
    for _, row in df.iterrows():
        ws.append(list(row))
    wb.save(ARCHIVO)

try:
    df = cargar_datos()
except FileNotFoundError:
    st.error(f"⚠️ No se encontró el archivo `{ARCHIVO}`. Verifica que la carpeta `data/` exista.")
    st.stop()

# ─────────────────────────────────────────────
# TABS PRINCIPALES
# ─────────────────────────────────────────────
tab_agregar, tab_editar, tab_tabla = st.tabs(["➕ Agregar mes", "✏️ Editar mes existente", "📊 Ver tabla completa"])

# ══════════════════════════════════════════════
# TAB 1 — AGREGAR NUEVO MES
# ══════════════════════════════════════════════
with tab_agregar:
    st.markdown('<div class="section-title">📅 Nuevo período</div>', unsafe_allow_html=True)

    # Sugerir el mes siguiente al último registrado
    ultimo = df["mes_Proc"].dropna().iloc[-1] if not df.empty else "2026-01"
    try:
        dt_ultimo = datetime.strptime(str(ultimo)[:7], "%Y-%m")
        if dt_ultimo.month == 12:
            sugerido = f"{dt_ultimo.year + 1}-01"
        else:
            sugerido = f"{dt_ultimo.year}-{dt_ultimo.month + 1:02d}"
    except Exception:
        sugerido = ""

    col_a, col_b = st.columns([1, 2])
    with col_a:
        nuevo_mes = st.text_input("Mes de proceso", value=sugerido, placeholder="aaaa-mm",
                                  help="Formato: 2026-05")

    # Validar mes
    mes_valido = False
    if nuevo_mes:
        try:
            dt = datetime.strptime(nuevo_mes[:7], "%Y-%m")
            ultimo_dia = calendar.monthrange(dt.year, dt.month)[1]
            mes_valido = True
            if nuevo_mes in df["mes_Proc"].values:
                st.warning(f"⚠️ El mes **{nuevo_mes}** ya existe. Usa la pestaña **Editar** para modificarlo.")
                mes_valido = False
        except ValueError:
            st.error("Formato de mes inválido. Usa aaaa-mm (ej: 2026-06)")
            ultimo_dia = 30

    if mes_valido:
        st.markdown('<div class="section-title">💰 Valores del período</div>', unsafe_allow_html=True)

        # Tomar valores del último mes como referencia
        ultimo_row = df.iloc[-1] if not df.empty else {}

        def val_ref(col):
            try:
                v = ultimo_row.get(col, 0)
                return float(v) if pd.notna(v) else 0.0
            except Exception:
                return 0.0

        nuevo = {"mes_Proc": nuevo_mes}

        col1, col2, col3 = st.columns(3)
        campos = [c for c in LABELS if c not in ("mes_Proc", "Formato Fecha", "factor_sis")]

        for i, col in enumerate(campos):
            col_dest = [col1, col2, col3][i % 3]
            with col_dest:
                nuevo[col] = st.number_input(
                    LABELS[col],
                    value=val_ref(col),
                    format="%.4f" if col in ("sis", "Aporte AFP", "Seg Social Exp vida", "aporte_Ccaf", "aporte_Fonasa") else "%.2f",
                    key=f"new_{col}"
                )

        # factor_sis se calcula automáticamente
        sis_val = nuevo.get("sis", 0)
        nuevo["factor_sis"] = round(sis_val / 100, 6)
        nuevo["Formato Fecha"] = f"{ultimo_dia:02d}/{dt.month:02d}/{dt.year}"
        nuevo["ult_Diames"] = ultimo_dia

        st.markdown(f"📌 `factor_sis` calculado automáticamente: **{nuevo['factor_sis']}** &nbsp; | &nbsp; Último día del mes: **{ultimo_dia}**",
                    unsafe_allow_html=True)

        st.markdown("")
        if st.button("💾 Guardar nuevo mes", key="btn_guardar_nuevo"):
            fila = {col: nuevo.get(col, None) for col in df.columns}
            df_nuevo = pd.concat([df, pd.DataFrame([fila])], ignore_index=True)
            try:
                guardar_datos(df_nuevo)
                st.cache_data.clear()
                st.success(f"✅ Mes **{nuevo_mes}** agregado correctamente.")
                st.rerun()
            except Exception as e:
                st.error(f"Error al guardar: {e}")

# ══════════════════════════════════════════════
# TAB 2 — EDITAR MES EXISTENTE
# ══════════════════════════════════════════════
with tab_editar:
    st.markdown('<div class="section-title">🔍 Seleccionar período</div>', unsafe_allow_html=True)

    meses_disponibles = df["mes_Proc"].dropna().tolist()
    mes_sel = st.selectbox("Mes a editar", options=list(reversed(meses_disponibles)),
                           help="Los meses más recientes aparecen primero")

    if mes_sel:
        idx = df[df["mes_Proc"] == mes_sel].index[0]
        fila_actual = df.loc[idx].copy()

        st.markdown('<div class="section-title">✏️ Editar valores</div>', unsafe_allow_html=True)

        editado = {"mes_Proc": mes_sel}
        campos_e = [c for c in LABELS if c not in ("mes_Proc", "Formato Fecha", "factor_sis")]

        col1, col2, col3 = st.columns(3)
        for i, col in enumerate(campos_e):
            col_dest = [col1, col2, col3][i % 3]
            with col_dest:
                val_actual = fila_actual.get(col, 0)
                try:
                    val_actual = float(val_actual) if pd.notna(val_actual) else 0.0
                except Exception:
                    val_actual = 0.0
                editado[col] = st.number_input(
                    LABELS[col],
                    value=val_actual,
                    format="%.4f" if col in ("sis", "Aporte AFP", "Seg Social Exp vida", "aporte_Ccaf", "aporte_Fonasa") else "%.2f",
                    key=f"edit_{col}"
                )

        sis_edit = editado.get("sis", 0)
        editado["factor_sis"] = round(sis_edit / 100, 6)

        try:
            dt_e = datetime.strptime(mes_sel[:7], "%Y-%m")
            ult_dia_e = calendar.monthrange(dt_e.year, dt_e.month)[1]
            editado["Formato Fecha"] = f"{ult_dia_e:02d}/{dt_e.month:02d}/{dt_e.year}"
            editado["ult_Diames"] = ult_dia_e
        except Exception:
            pass

        st.markdown(f"📌 `factor_sis` calculado automáticamente: **{editado['factor_sis']}**")

        st.markdown("")
        if st.button("💾 Guardar cambios", key="btn_guardar_edit"):
            for col in df.columns:
                df.at[idx, col] = editado.get(col, df.at[idx, col])
            try:
                guardar_datos(df)
                st.cache_data.clear()
                st.success(f"✅ Mes **{mes_sel}** actualizado correctamente.")
                st.rerun()
            except Exception as e:
                st.error(f"Error al guardar: {e}")

# ══════════════════════════════════════════════
# TAB 3 — VER TABLA COMPLETA
# ══════════════════════════════════════════════
with tab_tabla:
    st.markdown('<div class="section-title">📊 Todos los períodos registrados</div>', unsafe_allow_html=True)

    col_f1, col_f2 = st.columns([1, 3])
    with col_f1:
        buscar = st.text_input("🔍 Filtrar por año", placeholder="ej: 2026")

    df_vista = df.copy()
    if buscar:
        df_vista = df_vista[df_vista["mes_Proc"].str.startswith(buscar)]

    # Renombrar columnas para mostrar
    df_display = df_vista.rename(columns=LABELS)

    st.dataframe(
        df_display,
        use_container_width=True,
        hide_index=True,
        height=500
    )

    st.markdown(f"<small style='color:#64748b'>Total: {len(df_vista)} períodos registrados</small>", unsafe_allow_html=True)
