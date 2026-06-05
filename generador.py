import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from io import BytesIO


def generar_archivo_entrada(file_conceptos, file_empresas):
    import streamlit as st
    import requests as req

    SUPABASE_URL = st.secrets['SUPABASE_URL']
    SUPABASE_KEY = st.secrets['SUPABASE_KEY']
    hdrs = {'apikey': SUPABASE_KEY, 'Authorization': f'Bearer {SUPABASE_KEY}'}

    def sb_list(table, col):
        r = req.get(f'{SUPABASE_URL}/rest/v1/{table}?select={col}', headers=hdrs)
        return [row[col] for row in r.json() if row.get(col)]

    # ── Leer archivos ─────────────────────────────────────────────────────────
    conceptos_df = pd.read_excel(file_conceptos, header=1)
    empresas_df  = pd.read_excel(file_empresas, header=1)

    # ── Leer instituciones desde Supabase ─────────────────────────────────────
    nombres_afp    = sb_list('inst_afp', 'nombre_afp')
    nombres_salud  = sb_list('inst_salud', 'nombre_inst')
    nombres_mutual = sb_list('inst_mutuales', 'nombre_institucion')
    nombres_caja   = sb_list('inst_cajas', 'nombre_institucion')

    # ── Extraer listas ────────────────────────────────────────────────────────
    hab_afecto = conceptos_df[conceptos_df['Tipo']=='Haber afecto']['Nombre'].tolist()
    hab_exento = conceptos_df[conceptos_df['Tipo']=='Haber exento']['Nombre'].tolist()
    descuentos = conceptos_df[conceptos_df['Tipo']=='Descuento']['Nombre'].tolist()

    nombres_empresa = empresas_df['Nombre'].dropna().tolist()

    # ── Columnas fijas ────────────────────────────────────────────────────────
    FIXED = ['mes_Proceso', 'rut_trabajador', 'num_contrato', 'nombre_emp', 'id_empresa',
             'Dias Perm y Faltas', 'Dias Lic. Med.', 'Ult. Imp 30 dias']

    DESC_LEGALES = [
        'Cotizacion AFP', 'id_afp', 'AFP Reliq meses anteriores', 'Ahorro AFP',
        'Cotizacion SALUD', 'id_salud', 'SALUD Reliq meses anteriores',
        'Seguro de Cesantia', 'CESANTIA Reliq meses anteriores',
        'Trabajo Pesado Empleado', 'APVI Ahorro voluntario mensual',
        'APVC Ahorro voluntario colectivo', 'APVI Deposito Convenido',
        'Afiliado Voluntario Cotizacion', 'Afiliado Voluntario Ahorro',
        'Prestamo solidario Remuneracion', 'Impuesto mensual',
        'IMPUESTO Reliq meses anteriores', 'Impuesto Agricola',
        'Mayor retencion Solicitada', 'Trabajo pesado Empl Reliq anteriores'
    ]

    APORTES_EMP = [
        'Trabajo Pesado', 'Trabajo pesado Reliq anteriores',
        'APVC - Aporte Empleador', 'Aporte a CCAF', 'id_ccaf',
        'CCAF Reliq meses anteriores', 'Aporte Seguro Salud',
        'Mutual', 'id_mutual', 'MUTUAL Reliq meses anteriores',
        'Seguro Invalidez y Sobrevivencia', 'SIS Reliq meses anteriores',
        'Seguro de Cesantia CI', 'CESANTIA CI Reliq meses anteriores',
        'Seguro de Cesantia Solidario', 'CESANTIA SOL Reliq meses anteriores',
        'Aporte AFP Empleador', 'Aporte AFP Empleador Reliq meses anteriores',
        'Aporte FAPP Compensación Expectativa de Vida', 'Aporte por Seguro Covid'
    ]

    all_cols = FIXED + hab_afecto + hab_exento + DESC_LEGALES + descuentos + APORTES_EMP

    # ── Meses 2020-01 a 2050-12 ───────────────────────────────────────────────
    meses = [f"{y}-{m:02d}" for y in range(2020, 2051) for m in range(1, 13)]

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liquidaciones"

    # ── Hojas ocultas ─────────────────────────────────────────────────────────
    def hidden_sheet(name, values, col2=None):
        sh = wb.create_sheet(name)
        for i, v in enumerate(values, 1):
            sh.cell(row=i, column=1, value=v)
            if col2:
                sh.cell(row=i, column=2, value=col2[i-1])
        sh.sheet_state = 'hidden'
        return sh

    hidden_sheet("lst_meses", meses)
    # empresas: col A = Nombre, col B = Empresa(id)
    sh_emp = wb.create_sheet("lst_empresas")
    for i, row in empresas_df.iterrows():
        sh_emp.cell(row=i+1, column=1, value=row['Nombre'])
        sh_emp.cell(row=i+1, column=2, value=row['Empresa'])
    sh_emp.sheet_state = 'hidden'
    n_emp = len(empresas_df)

    hidden_sheet("lst_afp", nombres_afp)
    hidden_sheet("lst_salud", nombres_salud)
    hidden_sheet("lst_mutual", nombres_mutual)
    hidden_sheet("lst_caja", nombres_caja)

    ws = wb["Liquidaciones"]

    # ── Estilos ───────────────────────────────────────────────────────────────
    FILLS = {
        'fixed':   PatternFill('solid', start_color='2E4057'),
        'hafecto': PatternFill('solid', start_color='1B6CA8'),
        'hexento': PatternFill('solid', start_color='2E8B57'),
        'dlegal':  PatternFill('solid', start_color='8B4513'),
        'desc':    PatternFill('solid', start_color='B8522A'),
        'aporte':  PatternFill('solid', start_color='6A0DAD'),
        'id':      PatternFill('solid', start_color='444444'),
    }
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
    align_c = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ID_COLS = {'id_afp', 'id_salud', 'id_ccaf', 'id_mutual', 'id_empresa'}

    def get_fill(col):
        if col in ID_COLS: return FILLS['id']
        if col in FIXED:   return FILLS['fixed']
        if col in hab_afecto: return FILLS['hafecto']
        if col in hab_exento: return FILLS['hexento']
        if col in DESC_LEGALES: return FILLS['dlegal']
        if col in descuentos:   return FILLS['desc']
        if col in APORTES_EMP:  return FILLS['aporte']
        return FILLS['fixed']

    col_map = {name: idx+1 for idx, name in enumerate(all_cols)}

    def cl(name): return get_column_letter(col_map[name])

    # ── Encabezados ───────────────────────────────────────────────────────────
    for idx, col in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=idx, value=col)
        cell.font = header_font
        cell.fill = get_fill(col)
        cell.alignment = align_c
        ws.column_dimensions[get_column_letter(idx)].width = 16

    ws.row_dimensions[1].height = 38
    ws.freeze_panes = 'A2'

    DATA_ROWS = 1000

    # ── Fórmula id_empresa ────────────────────────────────────────────────────
    emp_col = cl('nombre_emp')
    id_emp_idx = col_map['id_empresa']
    for row in range(2, DATA_ROWS + 1):
        ws.cell(row=row, column=id_emp_idx).value = (
            f'=IFERROR(VLOOKUP({emp_col}{row},lst_empresas!$A$1:$B${n_emp},2,0),"")'
        )

    # ── Validaciones ──────────────────────────────────────────────────────────
    def add_dv(ws, dv, col_name):
        dv.sqref = f"{cl(col_name)}2:{cl(col_name)}{DATA_ROWS}"
        ws.add_data_validation(dv)

    add_dv(ws, DataValidation(type="list", formula1=f"lst_meses!$A$1:$A${len(meses)}", allow_blank=True,
                               showErrorMessage=True, error="Seleccione período válido (ej: 2024-01)", errorTitle="Período inválido"), 'mes_Proceso')

    # Validación RUT: formato XXXXXXXX-X y dígito verificador
    rut_col = cl('rut_trabajador')
    rut_formula = (
        f'=AND('
        f'LEN({rut_col}2)>=9,'
        f'LEN({rut_col}2)<=10,'
        f'MID({rut_col}2,LEN({rut_col}2)-1,1)="-",'
        f'ISNUMBER(VALUE(LEFT({rut_col}2,LEN({rut_col}2)-2))),'
        f'OR(RIGHT({rut_col}2,1)="K",'
        f'ISNUMBER(VALUE(RIGHT({rut_col}2,1)))),'
        f'MOD(11-MOD('
        f'VALUE(MID({rut_col}2,1,1))*3+'
        f'IF(LEN({rut_col}2)-2>=2,VALUE(MID({rut_col}2,2,1))*2,0)+'
        f'IF(LEN({rut_col}2)-2>=3,VALUE(MID({rut_col}2,3,1))*7,0)+'
        f'IF(LEN({rut_col}2)-2>=4,VALUE(MID({rut_col}2,4,1))*6,0)+'
        f'IF(LEN({rut_col}2)-2>=5,VALUE(MID({rut_col}2,5,1))*5,0)+'
        f'IF(LEN({rut_col}2)-2>=6,VALUE(MID({rut_col}2,6,1))*4,0)+'
        f'IF(LEN({rut_col}2)-2>=7,VALUE(MID({rut_col}2,7,1))*3,0)+'
        f'IF(LEN({rut_col}2)-2>=8,VALUE(MID({rut_col}2,8,1))*2,0),'
        f'11),11)=IF(RIGHT({rut_col}2,1)="K",11-10,VALUE(RIGHT({rut_col}2,1))))'
    )
    dv_rut = DataValidation(
        type="custom", formula1=rut_formula, allow_blank=True,
        showErrorMessage=True,
        error="RUT inválido. Verifique el formato (ej: 12345678-9) y el dígito verificador.",
        errorTitle="RUT inválido"
    )
    add_dv(ws, dv_rut, 'rut_trabajador')

    add_dv(ws, DataValidation(type="whole", operator="greaterThan", formula1="0",
                               allow_blank=True, showErrorMessage=True, error="Debe ser número entero mayor que 0", errorTitle="Contrato inválido"), 'num_contrato')

    add_dv(ws, DataValidation(type="list", formula1=f"lst_empresas!$A$1:$A${n_emp}", allow_blank=True,
                               showErrorMessage=True, error="Seleccione empresa de la lista", errorTitle="Empresa inválida"), 'nombre_emp')

    add_dv(ws, DataValidation(type="list", formula1=f"lst_afp!$A$1:$A${len(nombres_afp)}", allow_blank=True,
                               showErrorMessage=True, error="Seleccione AFP de la lista", errorTitle="AFP inválida"), 'id_afp')

    add_dv(ws, DataValidation(type="list", formula1=f"lst_salud!$A$1:$A${len(nombres_salud)}", allow_blank=True,
                               showErrorMessage=True, error="Seleccione institución de salud", errorTitle="Salud inválida"), 'id_salud')

    add_dv(ws, DataValidation(type="list", formula1=f"lst_mutual!$A$1:$A${len(nombres_mutual)}", allow_blank=True,
                               showErrorMessage=True, error="Seleccione mutual", errorTitle="Mutual inválida"), 'id_mutual')

    add_dv(ws, DataValidation(type="list", formula1=f"lst_caja!$A$1:$A${len(nombres_caja)}", allow_blank=True,
                               showErrorMessage=True, error="Seleccione caja de compensación", errorTitle="CCAF inválida"), 'id_ccaf')

    # Validación para columnas de días
    for col in ['Dias Perm y Faltas', 'Dias Lic. Med.', 'Ult. Imp 30 dias']:
        dv = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0",
                            allow_blank=True, showErrorMessage=True,
                            error="Debe ser número entero mayor o igual a 0", errorTitle="Valor inválido")
        dv.sqref = f"{cl(col)}2:{cl(col)}{DATA_ROWS}"
        ws.add_data_validation(dv)

    # Advertencia: si Dias Lic. Med. > 0, Ult. Imp 30 dias debe ser > 0
    lic_col     = cl('Dias Lic. Med.')
    ult_imp_col = cl('Ult. Imp 30 dias')
    dv_ult_imp = DataValidation(
        type="custom",
        formula1=f'=OR({lic_col}2=0,{lic_col}2="",(AND({lic_col}2>0,{ult_imp_col}2>0)))',
        allow_blank=True,
        showErrorMessage=True,
        errorStyle="warning",
        error='Si "Dias Lic. Med." es mayor que 0, el campo "Ult. Imp 30 dias" también debe ser mayor que 0.',
        errorTitle="Advertencia: Ult. Imp 30 dias"
    )
    dv_ult_imp.sqref = f"{ult_imp_col}2:{ult_imp_col}{DATA_ROWS}"
    ws.add_data_validation(dv_ult_imp)

    # Validación numérica para columnas de montos
    skip = {'mes_Proceso','rut_trabajador','num_contrato','nombre_emp','id_empresa',
            'id_afp','id_salud','id_mutual','id_ccaf',
            'Dias Perm y Faltas','Dias Lic. Med.','Ult. Imp 30 dias'}
    for col in all_cols:
        if col not in skip:
            dv = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True, showErrorMessage=False)
            dv.sqref = f"{cl(col)}2:{cl(col)}{DATA_ROWS}"
            ws.add_data_validation(dv)

    # ── Guardar en memoria ────────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
