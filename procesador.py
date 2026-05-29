import pandas as pd
import numpy as np
from io import BytesIO
import calendar


def es_bisiesto(anio):
    return calendar.isleap(int(anio))


def dias_mes(mes_proc):
    partes = str(mes_proc).split('-')
    anio, mes = int(partes[0]), int(partes[1])
    if mes in [1, 3, 5, 7, 8, 10, 12]:
        return 31
    elif mes in [4, 6, 9, 11]:
        return 30
    elif mes == 2:
        return 29 if es_bisiesto(anio) else 28
    return 30


def safe_float(val):
    try:
        if pd.isna(val):
            return 0.0
        return float(val)
    except:
        return 0.0


def procesar_liquidaciones(file_entrada, file_empleados, file_empresas, file_conceptos=None):
    # ── Leer archivos de referencia ───────────────────────────────────────────
    import streamlit as st
    import requests as req

    SUPABASE_URL = st.secrets['SUPABASE_URL']
    SUPABASE_KEY = st.secrets['SUPABASE_KEY']
    hdrs = {'apikey': SUPABASE_KEY, 'Authorization': f'Bearer {SUPABASE_KEY}'}

    def sb_df(table):
        r = req.get(f'{SUPABASE_URL}/rest/v1/{table}?select=*', headers=hdrs)
        return pd.DataFrame(r.json())

    afp_df      = sb_df('inst_afp').set_index('nombre_afp')
    salud_df    = sb_df('inst_salud').set_index('nombre_inst')
    mutuales_df = sb_df('inst_mutuales').set_index('nombre_institucion')
    cajas_df    = sb_df('inst_cajas').set_index('nombre_institucion')
    params_df   = sb_df('parametros_mensuales').set_index('mes_proc')
    params_df.index = params_df.index.astype(str)

    empresas_df = pd.read_excel(file_empresas, header=1).set_index('Nombre')
    entrada_df  = pd.read_excel(file_entrada)

    # ── Mapa Nombre → Concepto(ID) y Nombre → Tipo desde lista de conceptos ──
    nombre_a_id   = {}  # Nombre → Concepto (ID)
    nombre_a_tipo = {}  # Nombre → Tipo
    hab_exento_nombres = set()
    hab_afecto_nombres = set()

    if file_conceptos is not None:
        conc_df = pd.read_excel(file_conceptos, header=1)
        for _, r in conc_df.iterrows():
            nombre_a_id[str(r['Nombre'])]   = str(r['Concepto'])
            nombre_a_tipo[str(r['Nombre'])] = str(r['Tipo'])
            if r['Tipo'] == 'Haber exento':
                hab_exento_nombres.add(str(r['Nombre']))
            if r['Tipo'] == 'Haber afecto':
                hab_afecto_nombres.add(str(r['Nombre']))

    # ── Columnas fijas del archivo de entrada ─────────────────────────────────
    COLS_FIJAS = {'mes_Proceso', 'rut_trabajador', 'num_contrato', 'nombre_emp',
                  'id_empresa', 'id_afp', 'id_salud', 'id_mutual', 'id_ccaf'}
    concepto_cols = [c for c in entrada_df.columns if c not in COLS_FIJAS]

    # ── Mapeos por Nombre de columna ──────────────────────────────────────────
    CONCEPTOS_AFP_INST = {
        'Cotizacion AFP', 'AFP Reliq meses anteriores', 'Ahorro AFP',
        'Seguro de Cesantia', 'CESANTIA Reliq meses anteriores',
        'Trabajo Pesado Empleado', 'APVI Ahorro voluntario mensual',
        'APVC Ahorro voluntario colectivo', 'APVI Deposito Convenido',
        'Afiliado Voluntario Cotizacion', 'Afiliado Voluntario Ahorro',
        'Trabajo pesado Empl Reliq anteriores', 'Seguro Invalidez y Sobrevivencia',
        'SIS Reliq meses anteriores', 'CESANTIA CI Reliq meses anteriores',
        'CESANTIA SOL Reliq meses anteriores', 'Seguro de Cesantia CI',
        'Aporte AFP Empleador', 'Aporte AFP Empleador Reliq meses anteriores',
        'Aporte FAPP Compensación Expectativa de Vida',
        'Seguro de Cesantia Solidario', 'Trabajo Pesado',
        'Trabajo pesado Reliq anteriores', 'APVC - Aporte Empleador'
    }
    CONCEPTOS_SALUD_INST  = {'Cotizacion SALUD', 'SALUD Reliq meses anteriores', 'Aporte Seguro Salud'}
    CONCEPTOS_MUTUAL_INST = {'Mutual', 'MUTUAL Reliq meses anteriores'}
    CONCEPTOS_CCAF_INST   = {'Aporte a CCAF', 'CCAF Reliq meses anteriores'}

    CONCEPTOS_AFP_AFECTO = {
        'Cotizacion AFP', 'AFP Reliq meses anteriores', 'Ahorro AFP',
        'Trabajo Pesado', 'Trabajo pesado Reliq anteriores',
        'Trabajo Pesado Empleado', 'Trabajo pesado Empl Reliq anteriores',
        'Mutual', 'MUTUAL Reliq meses anteriores',
        'Seguro Invalidez y Sobrevivencia', 'SIS Reliq meses anteriores',
        'Aporte AFP Empleador', 'Aporte AFP Empleador Reliq meses anteriores',
        'Aporte FAPP Compensación Expectativa de Vida', 'APVC - Aporte Empleador'
    }
    CONCEPTOS_CES_AFECTO = {
        'Seguro de Cesantia', 'CESANTIA Reliq meses anteriores',
        'Seguro de Cesantia CI', 'CESANTIA CI Reliq meses anteriores',
        'Seguro de Cesantia Solidario', 'CESANTIA SOL Reliq meses anteriores',
    }

    # Índice donde empiezan los descuentos legales
    cotizacion_afp_idx = concepto_cols.index('Cotizacion AFP') if 'Cotizacion AFP' in concepto_cols else len(concepto_cols)
    haberes_cols = concepto_cols[:cotizacion_afp_idx]

    filas_salida = []

    for _, row in entrada_df.iterrows():
        mes_proc      = str(row.get('mes_Proceso', ''))
        rut           = row.get('rut_trabajador', '')
        num_cont      = row.get('num_contrato', '')
        nombre_emp    = row.get('nombre_emp', '')
        id_empresa    = row.get('id_empresa', '')
        id_afp_val    = str(row.get('id_afp', ''))
        id_salud_val  = str(row.get('id_salud', ''))
        id_mutual_val = str(row.get('id_mutual', ''))
        id_ccaf_val   = str(row.get('id_ccaf', ''))

        # Parámetros del mes
        params_row = params_df.loc[mes_proc] if mes_proc in params_df.index else None
        tope_afp   = safe_float(params_row['topeImp_pesos_afp']) if params_row is not None else 0
        tope_ces   = safe_float(params_row['topeCes_pesos'])     if params_row is not None else 0
        tope_salud = safe_float(params_row['topeSalud_pesos'])   if params_row is not None else 0

        # Totales haberes
        total_hab_afecto = sum(safe_float(row.get(c, 0)) for c in haberes_cols if c in hab_afecto_nombres or c not in hab_exento_nombres)
        total_hab_exento = sum(safe_float(row.get(c, 0)) for c in haberes_cols if c in hab_exento_nombres)

        # Valores clave
        val_afp            = safe_float(row.get('Cotizacion AFP', 0))
        val_ces_empleado   = safe_float(row.get('Seguro de Cesantia', 0))
        val_trab_pesa_empl = safe_float(row.get('Trabajo Pesado Empleado', 0))
        val_isapre         = safe_float(row.get('Cotizacion SALUD', 0))
        val_sueldo_base    = safe_float(row.get('Sueldo Base', 0))
        val_licencia       = safe_float(row.get('licenciaDias', 0)) if 'licenciaDias' in row.index else 0

        # Días
        dias_lic   = val_licencia if val_licencia > 0 else 0
        total_dias = dias_mes(mes_proc)
        dias_trab  = total_dias - dias_lic
        monto_init = (val_sueldo_base / dias_trab * 30) if dias_trab > 0 else 0

        # Helpers lookup
        def lookup_param(col):
            if params_row is not None and col in params_df.columns:
                return safe_float(params_row[col])
            return 0

        def lookup_afp_id():
            return afp_df.loc[id_afp_val, 'id_afp'] if id_afp_val in afp_df.index else ''

        def lookup_salud_id():
            return salud_df.loc[id_salud_val, 'id_inst'] if id_salud_val in salud_df.index else ''

        def lookup_mutual_id():
            return mutuales_df.loc[id_mutual_val, 'id_institucion'] if id_mutual_val in mutuales_df.index else ''

        def lookup_ccaf_id():
            return cajas_df.loc[id_ccaf_val, 'id_institucion'] if id_ccaf_val in cajas_df.index else ''

        # ── Por cada concepto ─────────────────────────────────────────────────
        for concepto_nombre in concepto_cols:
            monto = safe_float(row.get(concepto_nombre, 0))
            if monto <= 0:
                continue

            # Col 4 — Id del concepto (usar ID, no Nombre)
            id_concepto = nombre_a_id.get(concepto_nombre, concepto_nombre)

            # Col 5 — Monto del concepto
            monto_concepto = monto

            # Col 6 — Afecto
            if concepto_nombre in {'totalesEmpl', 'Sueldo liquido'}:
                afecto = total_hab_afecto + total_hab_exento
            elif concepto_nombre in CONCEPTOS_AFP_AFECTO:
                afecto = min(total_hab_afecto, tope_afp)
            elif concepto_nombre in CONCEPTOS_CES_AFECTO:
                afecto = min(total_hab_afecto, tope_ces)
            elif concepto_nombre in {'Impuesto mensual', 'IMPUESTO Reliq meses anteriores'}:
                afecto = total_hab_afecto - (val_afp + val_ces_empleado + val_trab_pesa_empl + min(val_isapre, tope_salud))
            elif concepto_nombre in {'SALUD Reliq meses anteriores'}:
                afecto = ''
            else:
                afecto = 0

            # Col 7 — Id de institución
            if concepto_nombre in CONCEPTOS_AFP_INST:
                id_inst = lookup_afp_id()
            elif concepto_nombre in CONCEPTOS_SALUD_INST:
                id_inst = lookup_salud_id()
            elif concepto_nombre in CONCEPTOS_MUTUAL_INST:
                id_inst = lookup_mutual_id()
            elif concepto_nombre in CONCEPTOS_CCAF_INST:
                id_inst = lookup_ccaf_id()
            else:
                id_inst = ''

            # Col 8 — Cotización de jubilación
            if concepto_nombre == 'Cotizacion AFP':
                cot_jub = safe_float(afp_df.loc[id_afp_val, 'cot_afp']) if id_afp_val in afp_df.index else ''
            elif concepto_nombre == 'Cotizacion SALUD':
                cot_jub = monto_concepto
            elif concepto_nombre == 'Aporte a CCAF':
                cot_jub = lookup_param('aporte_Ccaf')
            elif concepto_nombre == 'Mutual':
                cot_jub = safe_float(empresas_df.loc[str(nombre_emp), 'Cotización Mutual']) if str(nombre_emp) in empresas_df.index else ''
            elif concepto_nombre == 'Seguro Invalidez y Sobrevivencia':
                cot_jub = lookup_param('sis')
            elif concepto_nombre == 'Aporte AFP Empleador':
                cot_jub = lookup_param('Aporte AFP')
            elif concepto_nombre == 'Aporte FAPP Compensación Expectativa de Vida':
                cot_jub = lookup_param('Seg Social Exp vida')
            else:
                cot_jub = ''

            # Col 9 — Días de licencias
            dias_licencias = monto if concepto_nombre == 'licenciaDias' and monto > 0 else 0

            # Col 10 — Días trabajados
            dias_trabajados = dias_trab

            # Col 13 — Total rebajas LLSS
            if concepto_nombre == 'Impuesto mensual':
                total_rebajas = val_afp + val_ces_empleado + val_trab_pesa_empl + min(val_isapre, tope_salud)
            else:
                total_rebajas = 0

            # Col 14 — Rentas no gravadas
            if concepto_nombre == 'Impuesto mensual':
                rentas_no_grav = total_hab_exento
            else:
                rentas_no_grav = 0

            filas_salida.append({
                'Fecha de proceso':          mes_proc,
                'Id empleado':               rut,
                'Número de contrato':        num_cont,
                'Id del concepto':           id_concepto,
                'Monto del concepto':        monto_concepto,
                'Afecto':                    afecto,
                'Id de institución':         id_inst,
                'Cotización de jubilación':  cot_jub,
                'Días de licencias':         dias_licencias,
                'Días trabajados':           dias_trabajados,
                'Fecha de aplicación':       'x',
                'Empresa':                   id_empresa,
                'Total de rebajas por LLSS': total_rebajas,
                'Rentas no gravadas':        rentas_no_grav,
                'Rebaja por zona extrema':   0,
                'Jornada':                   'C',
                'Días de vacaciones':        '',
                'Monto Init':                round(monto_init, 2),
                'Fase':                      1,
            })

    df_salida = pd.DataFrame(filas_salida)
    n_trabajadores = len(entrada_df)
    n_filas = len(df_salida)

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df_salida.to_excel(writer, index=False, sheet_name='Liquidaciones detalladas')
        ws = writer.sheets['Liquidaciones detalladas']
        from openpyxl.styles import Font, PatternFill, Alignment
        hfont  = Font(name='Arial', bold=True, color='FFFFFF', size=9)
        hfill  = PatternFill('solid', start_color='1a2744')
        halign = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for cell in ws[1]:
            cell.font  = hfont
            cell.fill  = hfill
            cell.alignment = halign
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = 'A2'
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

    buf.seek(0)
    return buf.read(), n_filas, n_trabajadores
