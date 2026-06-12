import streamlit as st
import pandas as pd
import numpy as np
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
import os
from datetime import datetime

# ─────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────

DATA_DIR = "data"

st.set_page_config(
    page_title="Rex+ | LRE a Detalle Liquidaciones",
    page_icon="💼",
    layout="wide"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
.rex-header { background-color: #1a2744; padding: 14px 28px; border-radius: 10px;
    display: flex; align-items: center; justify-content: space-between; margin-bottom: 28px; }
.rex-logo { background: white; color: #1a2744; font-weight: 800; font-size: 15px;
    padding: 5px 10px; border-radius: 6px; letter-spacing: 0.5px; }
.rex-logo span { color: #00b4d8; }
.rex-title { color: white; font-size: 18px; font-weight: 600; margin-left: 16px; }
.rex-badge { background: #00b4d8; color: white; font-size: 11px; font-weight: 700;
    padding: 4px 12px; border-radius: 20px; letter-spacing: 1px; }
.step-card { background: white; border: 1px solid #e8edf5; border-radius: 10px;
    padding: 18px 20px; margin-bottom: 12px; }
.step-label { color: #00b4d8; font-size: 11px; font-weight: 700; letter-spacing: 1px;
    text-transform: uppercase; margin-bottom: 4px; }
.step-title { font-size: 15px; font-weight: 600; color: #1a2744; margin-bottom: 4px; }
.step-desc { font-size: 13px; color: #6b7a9a; }
.section-title { font-size: 20px; font-weight: 700; color: #1a2744; margin-bottom: 4px; }
.section-sub { font-size: 13px; color: #6b7a9a; margin-bottom: 20px; }
.alert-error { background: #fff0f0; border-left: 4px solid #e53e3e; border-radius: 6px;
    padding: 12px 16px; margin: 8px 0; font-size: 13px; color: #c53030; }
.alert-success { background: #f0fff4; border-left: 4px solid #38a169; border-radius: 6px;
    padding: 12px 16px; margin: 8px 0; font-size: 13px; color: #276749; }
.alert-warning { background: #fffbf0; border-left: 4px solid #d69e2e; border-radius: 6px;
    padding: 12px 16px; margin: 8px 0; font-size: 13px; color: #744210; }
.rex-divider { border: none; border-top: 1px solid #e8edf5; margin: 24px 0; }
.stButton > button { background-color: #1a2744 !important; color: white !important;
    border: none !important; border-radius: 8px !important; font-weight: 600 !important;
    padding: 8px 20px !important; font-size: 14px !important; }
.stButton > button:hover { background-color: #00b4d8 !important; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────

COL_FECHA_PROCESO   = "Fecha de proceso"
COL_ID_EMPLEADO     = "Id empleado"
COL_NRO_CONTRATO    = "Número de contrato"
COL_ID_EMPRESA      = "Id de empresa"
COL_AFP             = "afp"
COL_PCT_AFP         = "%afp"
COL_ISAPRE          = "isapre"
COL_CCAF            = "Ccaf"
COL_PCT_CCAF        = "% Ccaf"
COL_MUTUAL          = "Mutual"
COL_PCT_MUTUAL      = "% mutual"
COL_PCT_SIS         = "%Sis"
COL_DIAS_TRABAJADOS = "Nro días trabajados"
COL_DIAS_LICENCIA   = "Nro días de licencia médica"

# Códigos de haberes afectos sin tope (para base de cálculo AFP/SIS/Mutual)
CODIGOS_HAB_AFECTOS = {
    "2101","2102","2103","2104","2105","2106","2107","2108","2110",
    "2111","2112","2113","2115","2123","2124","2161","2201","2202"
}

# Códigos para rentas no gravadas (base del campo cuando concepto = impuesto)
CODIGOS_RENTAS_NO_GRAVADAS = {
    "2204","2301","2302","2303","2304","2305","2306","2308","2309",
    "2310","2311","2312","2313","2314","2315","2316","2331","2417","2418"
}

CONCEPTOS_AFECTO_AFP = {"afp","isapre","trabajoPesaEmpl","trabajoPesa","sis","mutual"}
CONCEPTOS_AFECTO_CES = {"cesEmpleado","cesAporteCi","cesAporteSol"}
CONCEPTOS_SIEMPRE    = {"impuesto","cesEmpleado"}   # incluir aunque monto = 0

# Códigos para validación de cuadre
CODIGOS_VAL_HAB_AFECTOS = {
    "2101","2102","2103","2104","2105","2106","2107","2108","2110",
    "2111","2112","2113","2115","2123","2124","2201","2202"
}
CODIGOS_VAL_HAB_EXENTOS = {
    "2204","2301","2302","2303","2304","2305","2306","2308","2309",
    "2310","2311","2312","2313","2314","2315","2316","2331","2417","2418"
}
CODIGOS_VAL_DESC_LEGALES = {
    "3141","3143","3144","3151","3154","3155","3156",
    "3161","3162","3163","3164","3166"
}
CODIGOS_VAL_OTROS_DESC = {
    "3171","3110","3181","3182","3183","3185","3186","3188"
}
COD_TOTAL_LIQUIDO = "5501"

# ─────────────────────────────────────────────
# FUNCIONES AUXILIARES
# ─────────────────────────────────────────────

def extraer_codigo(nombre):
    """Extrae código numérico de 4+ dígitos entre paréntesis.
    Soporta: Sueldo(2101) y Viáticos (cód 2303)."""
    # Buscar todos los números de 4+ dígitos seguidos de )
    m = re.findall(r"(\d{4,})\)", str(nombre))
    return m[-1] if m else None

def round_int(val):
    """Convierte a entero redondeado, sin decimales."""
    if val == "" or val is None:
        return ""
    try:
        return int(round(float(val)))
    except:
        return 0

def parse_pct(val):
    """Convierte porcentaje con coma decimal ('11,44') a float."""
    if val is None:
        return 0.0
    try:
        return float(str(val).replace(",", "."))
    except Exception:
        return 0.0

def safe_num(val, default=0):
    """Convierte a float de forma segura."""
    if val is None:
        return default
    if isinstance(val, float) and np.isnan(val):
        return default
    try:
        return float(val)
    except Exception:
        return default

def get_str(row, col, default=""):
    """Lee una columna como string."""
    if col not in row.index:
        return default
    v = row[col]
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return default
    return str(v).strip()

def sum_codes(row, all_cols, codigos):
    """Suma columnas cuyo código numérico esté en el set 'codigos'."""
    total = 0.0
    for col in all_cols:
        if extraer_codigo(col) in codigos:
            total += safe_num(row.get(col, 0))
    return total

def get_param(df_params, mes_proc, campo):
    """Obtiene un parámetro mensual."""
    if df_params.empty:
        return 0
    r = df_params[df_params["mes_Proc"] == str(mes_proc)]
    if r.empty:
        return 0
    return safe_num(r.iloc[0].get(campo, 0))

def construir_mapeo(equiv_dict_raw, df_cols):
    """
    Construye mapeo columna_LRE → id_concepto usando el código numérico
    para tolerar diferencias de texto entre equiv_conceptos y el LRE real.
    """
    # Código → concepto desde equiv
    cod_to_concepto = {}
    for cod_lre, concepto in equiv_dict_raw.items():
        cod = extraer_codigo(cod_lre)
        if cod:
            cod_to_concepto[cod] = concepto

    # Columna LRE → concepto
    col_to_concepto = {}
    for col in df_cols:
        cod = extraer_codigo(col)
        if cod and cod in cod_to_concepto:
            col_to_concepto[col] = cod_to_concepto[cod]

    return col_to_concepto

# ─────────────────────────────────────────────
# CARGA DE REFERENCIAS (con caché Streamlit)
# ─────────────────────────────────────────────

@st.cache_data
def cargar_equiv_conceptos():
    path = os.path.join(DATA_DIR, "equiv_conceptos_ab_rex.xlsx")
    if os.path.exists(path):
        df = pd.read_excel(path)
        return dict(zip(df["cod_lre"], df["concepto_detalle"]))
    return {}

@st.cache_data
def cargar_parametros():
    path = os.path.join(DATA_DIR, "parametrosMesuales.xlsx")
    if os.path.exists(path):
        df = pd.read_excel(path, sheet_name="Hoja2", dtype={"mes_Proc": str})
        df["mes_Proc"] = df["mes_Proc"].astype(str).str.strip()
        return df
    return pd.DataFrame()

# ─────────────────────────────────────────────
# VALIDACIÓN DE CUADRE
# ─────────────────────────────────────────────

def validar_cuadre(df_entrada):
    """
    Valida que (Total Haberes Afectos + Total Haberes Exentos) 
    - (Total Descuentos Legales + Total Otros Descuentos) == Total Líquido(5501)
    Retorna (ok: bool, df_descuadres: DataFrame)
    """
    lre_cols = list(df_entrada.columns)

    registros = []
    for _, row in df_entrada.iterrows():
        hab_afectos   = sum_codes(row, lre_cols, CODIGOS_VAL_HAB_AFECTOS)
        hab_exentos   = sum_codes(row, lre_cols, CODIGOS_VAL_HAB_EXENTOS)
        desc_legales  = sum_codes(row, lre_cols, CODIGOS_VAL_DESC_LEGALES)
        otros_desc    = sum_codes(row, lre_cols, CODIGOS_VAL_OTROS_DESC)

        liquido_calc  = (hab_afectos + hab_exentos) - (desc_legales + otros_desc)

        # Buscar columna de total líquido por código 5501
        liquido_inf   = next((safe_num(row.get(c, 0)) for c in lre_cols if extraer_codigo(c) == COD_TOTAL_LIQUIDO), 0)
        diferencia    = round_int(liquido_calc) - round_int(liquido_inf)

        registros.append({
            "hab_afectos":  hab_afectos,
            "hab_exentos":  hab_exentos,
            "desc_legales": desc_legales,
            "otros_desc":   otros_desc,
            "liq_calc":     round_int(liquido_calc),
            "liq_inf":      round_int(liquido_inf),
            "diferencia":   diferencia,
            "descuadrado":  diferencia != 0,
        })

    df_val = pd.DataFrame(registros)
    df_entrada_r = df_entrada.reset_index(drop=True)
    mask = df_val["descuadrado"].values
    df_desc = df_entrada_r[mask].copy()

    if not df_desc.empty:
        df_val_desc = df_val[mask].reset_index(drop=True)
        df_desc = df_desc.reset_index(drop=True)
        df_desc["Total haberes afectos"]   = df_val_desc["hab_afectos"].values
        df_desc["Total haberes exentos"]   = df_val_desc["hab_exentos"].values
        df_desc["Total descuentos legales"]= df_val_desc["desc_legales"].values
        df_desc["Total otros descuentos"]  = df_val_desc["otros_desc"].values
        df_desc["Liquido calculado"]       = df_val_desc["liq_calc"].values
        df_desc["Liquido informado"]       = df_val_desc["liq_inf"].values
        df_desc["Diferencia"]              = df_val_desc["diferencia"].values

    todo_ok = df_desc.empty
    return todo_ok, df_desc

def generar_log_excel(df_desc):
    """Genera el archivo log de descuadres en Excel."""
    # Convertir columnas numéricas a int para evitar distorsiones al escribir
    df_desc = df_desc.copy()
    cols_extra = {"Total haberes afectos","Total haberes exentos",
                  "Total descuentos legales","Total otros descuentos",
                  "Liquido calculado","Liquido informado","Diferencia"}
    for col in df_desc.columns:
        if col in cols_extra:
            df_desc[col] = pd.to_numeric(df_desc[col], errors="coerce").fillna(0).astype(int)
        elif df_desc[col].dtype in [float, "float64"]:
            try:
                df_desc[col] = df_desc[col].fillna(0).astype(int)
            except Exception:
                pass
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Descuadres"

    header_fill    = PatternFill("solid", fgColor="1A2744")
    header_font    = Font(bold=True, color="FFFFFF", size=10)
    error_fill     = PatternFill("solid", fgColor="FFE0E0")
    error_font     = Font(bold=True, color="C53030", size=10)
    border = Border(
        bottom=Side(style="thin", color="E8EDF5"),
        right=Side(style="thin", color="E8EDF5")
    )

    cols_extra = {"Total haberes afectos","Total haberes exentos",
                  "Total descuentos legales","Total otros descuentos",
                  "Liquido calculado","Liquido informado","Diferencia"}

    cols = list(df_desc.columns)
    for ci, col in enumerate(cols, 1):
        is_extra = col in cols_extra
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = error_fill if is_extra else header_fill
        cell.font = error_font if is_extra else header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = max(len(col) + 4, 14)

    for ri, row_data in enumerate(df_desc.itertuples(index=False), 2):
        fill = PatternFill("solid", fgColor="FFF5F5") if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    wb.save(output)
    return output.getvalue()

# ─────────────────────────────────────────────
# TRANSFORMACIÓN PRINCIPAL
# ─────────────────────────────────────────────

def transformar_lre(df_entrada, equiv_dict_raw, df_params):
    """
    Transforma el DataFrame LRE al formato de salida (Migración).
    Una fila de entrada → N filas de salida (una por concepto con valor ≠ 0).
    """
    lre_cols = list(df_entrada.columns)
    col_to_concepto = construir_mapeo(equiv_dict_raw, lre_cols)
    filas_salida = []

    for _, row in df_entrada.iterrows():

        # ── Campos de identificación ──
        mes_proc    = get_str(row, COL_FECHA_PROCESO)
        id_empleado = get_str(row, COL_ID_EMPLEADO)
        nro_cont    = int(safe_num(row.get(COL_NRO_CONTRATO, 1), 1))
        id_empresa  = int(safe_num(row.get(COL_ID_EMPRESA, ""), 0)) if COL_ID_EMPRESA in row.index else ""
        afp_ent     = get_str(row, COL_AFP)
        pct_afp     = parse_pct(row[COL_PCT_AFP] if COL_PCT_AFP in row.index else 0)
        isapre_ent  = get_str(row, COL_ISAPRE)
        ccaf_ent    = get_str(row, COL_CCAF)
        pct_ccaf    = parse_pct(row[COL_PCT_CCAF] if COL_PCT_CCAF in row.index else 0)
        mutual_ent  = get_str(row, COL_MUTUAL)
        pct_mutual  = parse_pct(row[COL_PCT_MUTUAL] if COL_PCT_MUTUAL in row.index else 0)
        pct_sis     = parse_pct(row[COL_PCT_SIS] if COL_PCT_SIS in row.index else 0)
        dias_trab   = int(safe_num(row.get(COL_DIAS_TRABAJADOS, 0)))
        dias_lic    = int(safe_num(row.get(COL_DIAS_LICENCIA, 0)))

        # Sueldo base (código 2101) → Monto Init
        sueldo = next((safe_num(row.get(c, 0)) for c in lre_cols if extraer_codigo(c) == "2101"), 0.0)
        monto_init = round(sueldo / dias_trab * 30) if dias_trab > 0 else 0

        # ── Bases de cálculo ──
        total_hab_afectos = sum_codes(row, lre_cols, CODIGOS_HAB_AFECTOS)

        tope_afp   = get_param(df_params, mes_proc, "topeImp_pesos_afp")
        tope_ces   = get_param(df_params, mes_proc, "topeCes_pesos")
        tope_salud = get_param(df_params, mes_proc, "topeSalud_pesos")

        afecto_afp = min(total_hab_afectos, tope_afp) if tope_afp > 0 else total_hab_afectos
        afecto_ces = min(total_hab_afectos, tope_ces) if tope_ces > 0 else total_hab_afectos
        afecto_salud = min(total_hab_afectos, tope_salud) if tope_salud > 0 else total_hab_afectos

        # Isapre: monto = col(3143) + col(3144)
        monto_isapre = sum(safe_num(row.get(c, 0)) for c in lre_cols if extraer_codigo(c) in ("3143", "3144"))
        rebaja_llss_isapre = min(monto_isapre, tope_salud) if tope_salud > 0 else monto_isapre
        afecto_isapre = total_hab_afectos - rebaja_llss_isapre

        # Rentas no gravadas (para impuesto y totalesEmpl)
        rentas_no_gravadas = sum_codes(row, lre_cols, CODIGOS_RENTAS_NO_GRAVADAS)
        afecto_totales = total_hab_afectos + rentas_no_gravadas

        # ── Acumular montos por concepto ──
        conceptos_acumulados = {}
        for col_lre, id_concepto in col_to_concepto.items():
            monto = safe_num(row.get(col_lre, 0))
            conceptos_acumulados[id_concepto] = conceptos_acumulados.get(id_concepto, 0.0) + monto

        # ── Fila especial: licenciaDias ──
        if dias_lic > 0:
            filas_salida.append({
                "Fecha de proceso":          mes_proc,
                "Id empleado":               id_empleado,
                "Número de contrato":        nro_cont,
                "Id del concepto":           "licenciaDias",
                "Monto del concepto":        dias_lic,
                "Afecto":                    "",
                "Id de institución":         "",
                "Cotización de jubilación":  "",
                "Días de licencias":         dias_lic,
                "Días trabajados":           dias_trab,
                "Fecha de aplicación":       mes_proc,
                "Empresa":                   id_empresa,
                "Total de rebajas por LLSS": 0,
                "Rentas no gravadas":        0,
                "Rebaja por zona extrema":   0,
                "Jornada":                   "C",
                "Días de vacaciones":        0,
                "Monto Init":                round_int(monto_init),
                "Fase":                      1,
            })

        # ── Generar filas de salida ──
        for id_concepto, monto in conceptos_acumulados.items():

            # Isapre: monto especial
            if id_concepto == "isapre":
                monto = monto_isapre

            # Excluir conceptos con monto 0 (salvo los que siempre deben ir)
            if monto == 0 and id_concepto not in CONCEPTOS_SIEMPRE:
                continue

            # ── Id de institución ──
            if id_concepto in {"afp","cesEmpleado","trabajoPesaEmpl","cesAporteCi","cesAporteSol","trabajoPesa","sis"}:
                id_inst = afp_ent
            elif id_concepto == "isapre":
                id_inst = isapre_ent
            elif id_concepto == "cajaCred":
                id_inst = ccaf_ent
            elif id_concepto == "mutual":
                id_inst = mutual_ent
            else:
                id_inst = ""

            # ── Afecto ──
            if id_concepto in CONCEPTOS_AFECTO_AFP:
                afecto = round_int(afecto_afp)
            elif id_concepto in CONCEPTOS_AFECTO_CES:
                afecto = round_int(afecto_ces)
            elif id_concepto == "impuesto":
                afecto = round_int(total_hab_afectos - rebaja_llss_isapre)
            elif id_concepto == "totalesEmpl":
                afecto = round_int(afecto_totales)
            else:
                afecto = ""

            # ── Cotización de jubilación ──
            if id_concepto == "afp":
                cot_jub = pct_afp
            elif id_concepto == "isapre":
                cot_jub = round_int(monto_isapre)
            elif id_concepto == "cesEmpleado":
                cot_jub = 0.6
            elif id_concepto == "sis":
                cot_jub = pct_sis
            elif id_concepto == "cajaCred":
                cot_jub = pct_ccaf
            elif id_concepto == "mutual":
                cot_jub = pct_mutual
            else:
                cot_jub = ""

            # ── Total rebajas LLSS ──
            rebaja_llss = round_int(rebaja_llss_isapre) if id_concepto == "impuesto" else 0

            # ── Rentas no gravadas ──
            rng = round_int(rentas_no_gravadas) if id_concepto == "impuesto" else 0

            filas_salida.append({
                "Fecha de proceso":          mes_proc,
                "Id empleado":               id_empleado,
                "Número de contrato":        nro_cont,
                "Id del concepto":           id_concepto,
                "Monto del concepto":        round_int(monto),
                "Afecto":                    afecto,
                "Id de institución":         id_inst,
                "Cotización de jubilación":  cot_jub,
                "Días de licencias":         dias_lic,
                "Días trabajados":           dias_trab,
                "Fecha de aplicación":       mes_proc,
                "Empresa":                   id_empresa,
                "Total de rebajas por LLSS": rebaja_llss,
                "Rentas no gravadas":        rng,
                "Rebaja por zona extrema":   round(safe_num(next((row.get(c,0) for c in lre_cols if extraer_codigo(c)=="3167"), 0))) if id_concepto=="zonaExtrema" else 0,
                "Jornada":                   "C",
                "Días de vacaciones":        0,
                "Monto Init":                round_int(monto_init),
                "Fase":                      1,
            })

    return pd.DataFrame(filas_salida)

# ─────────────────────────────────────────────
# GENERADOR DE EXCEL
# ─────────────────────────────────────────────

def generar_excel(df_salida):
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Migración"

    header_fill = PatternFill("solid", fgColor="1A2744")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(
        bottom=Side(style="thin", color="E8EDF5"),
        right=Side(style="thin", color="E8EDF5")
    )

    cols = list(df_salida.columns)
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = max(len(col) + 4, 14)

    for ri, row_data in enumerate(df_salida.itertuples(index=False), 2):
        fill = PatternFill("solid", fgColor="EAF0F8") if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    wb.save(output)
    return output.getvalue()

# ─────────────────────────────────────────────
# INTERFAZ STREAMLIT
# ─────────────────────────────────────────────

st.markdown("""
<div class="rex-header">
  <div style="display:flex; align-items:center; gap:12px;">
    <div class="rex-logo">Rex<span>+</span></div>
    <span class="rex-title">LRE a Detalle Liquidaciones</span>
  </div>
  <div class="rex-badge">MIGRACIÓN</div>
</div>
""", unsafe_allow_html=True)

st.markdown('<div class="section-title">📂 Transformación LRE → Detalle de Liquidaciones</div>', unsafe_allow_html=True)
st.markdown('<div class="section-sub">Sube el archivo Excel LRE para generar el archivo de migración en el formato requerido.</div>', unsafe_allow_html=True)

col1, col2, col3 = st.columns(3)
with col1:
    st.markdown("""<div class="step-card">
        <div class="step-label">PASO 1</div>
        <div class="step-title">Subir archivo LRE</div>
        <div class="step-desc">Archivo Excel con datos de remuneraciones. Puede contener múltiples meses.</div>
    </div>""", unsafe_allow_html=True)
with col2:
    st.markdown("""<div class="step-card">
        <div class="step-label">PASO 2</div>
        <div class="step-title">Procesamiento automático</div>
        <div class="step-desc">Se aplican las reglas de mapeo y transformación por empleado y concepto.</div>
    </div>""", unsafe_allow_html=True)
with col3:
    st.markdown("""<div class="step-card">
        <div class="step-label">PASO 3</div>
        <div class="step-title">Descargar archivo</div>
        <div class="step-desc">Excel de salida listo para importar en el sistema destino.</div>
    </div>""", unsafe_allow_html=True)

st.markdown('<hr class="rex-divider">', unsafe_allow_html=True)

# Carga de referencias
equiv_dict = cargar_equiv_conceptos()
df_params  = cargar_parametros()

if not equiv_dict:
    st.markdown('<div class="alert-warning">⚠️ No se encontró <b>data/equiv_conceptos_ab_rex.xlsx</b>.</div>', unsafe_allow_html=True)
if df_params.empty:
    st.markdown('<div class="alert-warning">⚠️ No se encontró <b>data/parametrosMesuales.xlsx</b>.</div>', unsafe_allow_html=True)

# Upload
st.markdown("### 📤 Subir archivo LRE (Excel)")
archivo_lre = st.file_uploader(
    "Selecciona el archivo Excel de entrada (LRE)",
    type=["xlsx", "xls"],
    accept_multiple_files=False,
    help="Debe tener una hoja con datos de liquidaciones. Puede contener múltiples meses."
)

if archivo_lre:
    st.markdown(f'<div class="alert-success">✅ Archivo cargado: <b>{archivo_lre.name}</b></div>', unsafe_allow_html=True)

    try:
        df_entrada = pd.read_excel(archivo_lre, sheet_name=0, dtype=str)
        df_entrada[COL_FECHA_PROCESO] = df_entrada[COL_FECHA_PROCESO].astype(str).str.strip()

        # Convertir columnas numéricas (desde índice 14): limpiar puntos de miles y convertir a float
        for col in df_entrada.columns[14:]:
            if col == COL_FECHA_PROCESO:
                continue
            try:
                cleaned = df_entrada[col].astype(str).str.strip().str.replace(r"\.", "", regex=True).str.replace(",", ".", regex=False)
                numeric = pd.to_numeric(cleaned, errors="coerce")
                if numeric.notna().sum() > len(df_entrada) * 0.3:
                    df_entrada[col] = numeric.fillna(0)
            except Exception:
                pass
        n_meses = df_entrada[COL_FECHA_PROCESO].nunique() if COL_FECHA_PROCESO in df_entrada.columns else "N/D"
        st.caption(f"📊 {len(df_entrada):,} filas × {len(df_entrada.columns)} columnas | Meses: {n_meses}")
    except Exception as e:
        st.markdown(f'<div class="alert-error">❌ Error al leer el archivo: {e}</div>', unsafe_allow_html=True)
        st.stop()

    with st.expander("👁️ Vista previa (primeras 5 filas)"):
        st.dataframe(df_entrada.head(), use_container_width=True)

    # Filtro por mes
    if COL_FECHA_PROCESO in df_entrada.columns:
        meses_disponibles = sorted([str(m) for m in df_entrada[COL_FECHA_PROCESO].unique() if pd.notna(m) and str(m).strip() not in ("", "nan", "None")])
        meses_sel = st.multiselect(
            "Filtrar por mes(es) a procesar (vacío = todos)",
            options=meses_disponibles,
            default=[]
        )
        if meses_sel:
            df_entrada = df_entrada[df_entrada[COL_FECHA_PROCESO].isin(meses_sel)]
            st.caption(f"Procesando {len(df_entrada):,} filas para: {', '.join(meses_sel)}")

    if st.button("▶ Validar y generar archivo de salida"):
        # ── PASO 1: Validación de cuadre ──
        with st.spinner("Validando archivo de entrada..."):
            try:
                todo_ok, df_desc = validar_cuadre(df_entrada)
            except Exception as e:
                st.markdown(f'<div class="alert-error">❌ Error en validación: {e}</div>', unsafe_allow_html=True)
                import traceback
                st.code(traceback.format_exc())
                st.stop()

        if not todo_ok:
            st.markdown('<div class="alert-error">⚠️ El archivo presenta descuadres en sus valores. Se generará un archivo log con los registros descuadrados.</div>', unsafe_allow_html=True)
            st.caption(f"{len(df_desc):,} registro(s) con descuadre.")

            with st.expander("👁️ Registros con descuadre"):
                st.dataframe(df_desc[["Fecha de proceso","Id empleado","Número de contrato",
                                       "Total haberes afectos","Total haberes exentos",
                                       "Total descuentos legales","Total otros descuentos",
                                       "Liquido calculado","Liquido informado","Diferencia"]],
                             use_container_width=True)

            log_bytes = generar_log_excel(df_desc)
            nombre_base = os.path.splitext(archivo_lre.name)[0]
            nombre_log = f"log_descuadres_{nombre_base}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            st.download_button(
                label="⬇️ Descargar archivo log de descuadres (.xlsx)",
                data=log_bytes,
                file_name=nombre_log,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        else:
            st.markdown('<div class="alert-success">✅ Archivo validado exitosamente. Ahora puedes generar el archivo de salida.</div>', unsafe_allow_html=True)

            # ── PASO 2: Transformación ──
            with st.spinner("Generando archivo de salida..."):
                try:
                    df_salida = transformar_lre(df_entrada, equiv_dict, df_params)

                    if df_salida.empty:
                        st.markdown('<div class="alert-warning">⚠️ No se generaron filas de salida.</div>', unsafe_allow_html=True)
                    else:
                        st.markdown(f'<div class="alert-success">✅ Transformación completada: <b>{len(df_salida):,} filas</b> generadas.</div>', unsafe_allow_html=True)

                        with st.expander("👁️ Vista previa del resultado (primeras 20 filas)"):
                            st.dataframe(df_salida.head(20), use_container_width=True)

                        with st.expander("📊 Resumen por mes y concepto"):
                            resumen = (
                                df_salida.groupby(["Fecha de proceso", "Id del concepto"])["Monto del concepto"]
                                .sum().reset_index()
                            )
                            st.dataframe(resumen, use_container_width=True)

                        excel_bytes = generar_excel(df_salida)
                        nombre_base = os.path.splitext(archivo_lre.name)[0]
                        nombre_salida = f"migracion_{nombre_base}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

                        st.download_button(
                            label="⬇️ Descargar archivo de salida (.xlsx)",
                            data=excel_bytes,
                            file_name=nombre_salida,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                except Exception as e:
                    st.markdown(f'<div class="alert-error">❌ Error durante el procesamiento: {e}</div>', unsafe_allow_html=True)
                    import traceback
                    st.code(traceback.format_exc())
