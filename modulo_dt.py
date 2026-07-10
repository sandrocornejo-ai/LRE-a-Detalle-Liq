"""
modulo_dt.py
Procesamiento del archivo de la Dirección del Trabajo (DT)
para generación del archivo de liquidaciones en detalle Rex+.
"""

import re
import io
import os
import calendar
import pandas as pd
import numpy as np
import streamlit as st
from io import StringIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


# ─────────────────────────────────────────────
# CONSTANTES — códigos de columnas LRE (DT)
# ─────────────────────────────────────────────
COD_RUT           = 1101
COD_DIAS_TRAB     = 1115
COD_DIAS_LIC      = 1116
COD_DIAS_VAC      = 1117
COD_SUELDO        = 2101
COD_SALUD7        = 3143
COD_SALUD_VOL     = 3144
COD_AFP           = 3141
COD_CES_TRAB      = 3151
COD_APVI_MOD_B    = 3156
COD_TRAB_PESADO   = 3154
COD_REBAJA_ZONA   = 3167
COD_AFP_INST      = 1141
COD_ISAPRE_INST   = 1143
COD_MUTUAL_INST   = 1152
COD_CCAF_INST     = 1110

# ─────────────────────────────────────────────
# DICCIONARIO DE COLUMNAS LRE (DT)
# Fuente: Cod_Colum_Lre.xlsx — actualizar aquí si cambia
# ─────────────────────────────────────────────
LRE_COLUMNAS = {
    1101: "Rut trabajador (1101)",
    1102: "Fecha inicio contrato (1102)",
    1103: "Fecha término de contrato (1103)",
    1104: "Causal término de contrato (1104)",
    1105: "Región prestación de servicios (1105)",
    1106: "Comuna prestación de servicios (1106)",
    1170: "Tipo impuesto a la renta (1170)",
    1146: "Técnico extranjero exención cot. previsionales (1146)",
    1107: "Código tipo de jornada (1107)",
    1108: "Persona con Discapacidad - Pensionado por Invalidez (1108)",
    1109: "Pensionado por vejez (1109)",
    1141: "AFP (1141)",
    1142: "IPS (ExINP) (1142)",
    1143: "FONASA - ISAPRE (1143)",
    1151: "AFC (1151)",
    1110: "CCAF (1110)",
    1152: "Org. administrador ley 16.744 (1152)",
    1111: "Nro cargas familiares legales autorizadas (1111)",
    1112: "Nro de cargas familiares maternales (1112)",
    1113: "Nro de cargas familiares invalidez (1113)",
    1114: "Tramo asignación familiar (1114)",
    1171: "Rut org sindical 1 (1171)",
    1172: "Rut org sindical 2 (1172)",
    1173: "Rut org sindical 3 (1173)",
    1174: "Rut org sindical 4 (1174)",
    1175: "Rut org sindical 5 (1175)",
    1176: "Rut org sindical 6 (1176)",
    1177: "Rut org sindical 7 (1177)",
    1178: "Rut org sindical 8 (1178)",
    1179: "Rut org sindical 9 (1179)",
    1180: "Rut org sindical 10 (1180)",
    1115: "Nro días trabajados en el mes (1115)",
    1116: "Nro días de licencia médica en el mes (1116)",
    1117: "Nro días de vacaciones en el mes (1117)",
    1118: "Subsidio trabajador joven (1118)",
    1154: "Puesto Trabajo Pesado (1154)",
    1155: "APVI (1155)",
    1157: "APVC (1157)",
    1131: "Indemnización a todo evento (1131)",
    1132: "Tasa indemnización a todo evento (1132)",
    2101: "Sueldo (2101)",
    2102: "Sobresueldo (2102)",
    2103: "Comisiones (2103)",
    2104: "Semana corrida (2104)",
    2105: "Participación (2105)",
    2106: "Gratificación (2106)",
    2107: "Recargo 30% día domingo (2107)",
    2108: "Remun. variable pagada en vacaciones (2108)",
    2109: "Remun. variable pagada en clausura (2109)",
    2110: "Aguinaldo (2110)",
    2111: "Bonos u otras remun. fijas mensuales (2111)",
    2112: "Tratos (2112)",
    2113: "Bonos u otras remun. variables mensuales o superiores a un mes (2113)",
    2114: "Ejercicio opción no pactada en contrato (2114)",
    2115: "Beneficios en especie constitutivos de remun (2115)",
    2116: "Remuneraciones bimestrales (2116)",
    2117: "Remuneraciones trimestrales (2117)",
    2118: "Remuneraciones cuatrimestral (2118)",
    2119: "Remuneraciones semestrales (2119)",
    2120: "Remuneraciones anuales (2120)",
    2121: "Participación anual (2121)",
    2122: "Gratificación anual (2122)",
    2123: "Otras remuneraciones superiores a un mes (2123)",
    2124: "Pago por horas de trabajo sindical (2124)",
    2161: "Sueldo empresarial  (2161)",
    2201: "Subsidio por incapacidad laboral por licencia médica (2201)",
    2202: "Beca de estudio (2202)",
    2203: "Gratificaciones de zona (2203)",
    2204: "Otros ingresos no constitutivos de renta (2204)",
    2301: "Colación (2301)",
    2302: "Movilización (2302)",
    2303: "Viáticos (2303)",
    2304: "Asignación de pérdida de caja (2304)",
    2305: "Asignación de desgaste herramienta (2305)",
    2311: "Asignación familiar legal (2311)",
    2306: "Gastos por causa del trabajo (2306)",
    2307: "Gastos por cambio de residencia (2307)",
    2308: "Sala cuna (2308)",
    2309: "Asignación trabajo a distancia o teletrabajo (2309)",
    2347: "Depósito convenido hasta UF 900 (2347)",
    2310: "Alojamiento por razones de trabajo (2310)",
    2312: "Asignación de traslación (2312)",
    2313: "Indemnización por feriado legal (2313)",
    2314: "Indemnización años de servicio (2314)",
    2315: "Indemnización sustitutiva del aviso previo (2315)",
    2316: "Indemnización fuero maternal (2316)",
    2331: "Pago indemnización a todo evento (2331)",
    2417: "Indemnizaciones voluntarias tributables (2417)",
    2418: "Indemnizaciones contractuales tributables (2418)",
    3141: "Cotización obligatoria previsional (AFP o IPS) (3141)",
    3143: "Cotización obligatoria salud 7% (3143)",
    3144: "Cotización voluntaria para salud (3144)",
    3151: "Cotización AFC - trabajador (3151)",
    3146: "Cotizaciones técnico extranjero para seguridad social fuera de Chile (3146)",
    3147: "Descuento depósito convenido hasta UF 900 anual (3147)",
    3155: "Cotización APVi Mod A (3155)",
    3156: "Cotización APVi Mod B hasta UF50 (3156)",
    3157: "Cotización APVc Mod A (3157)",
    3158: "Cotización APVc Mod B hasta UF50 (3158)",
    3161: "Impuesto retenido por remuneraciones (3161)",
    3162: "Impuesto retenido por indemnizaciones (3162)",
    3163: "Mayor retención de impuestos solicitada por el trabajador (3163)",
    3164: "Impuesto retenido por reliquidación remun. devengadas otros períodos (3164)",
    3165: "Diferencia impuesto reliquidación remun. devengadas en este período (3165)",
    3166: "Retención préstamo clase media 2020 (Ley 21.252)  (3166)",
    3167: "Rebaja zona extrema DL 889  (3167)",
    3171: "Cuota sindical 1 (3171)",
    3172: "Cuota sindical 2 (3172)",
    3173: "Cuota sindical 3 (3173)",
    3174: "Cuota sindical 4 (3174)",
    3175: "Cuota sindical 5 (3175)",
    3176: "Cuota sindical 6 (3176)",
    3177: "Cuota sindical 7 (3177)",
    3178: "Cuota sindical 8 (3178)",
    3179: "Cuota sindical 9 (3179)",
    3180: "Cuota sindical 10 (3180)",
    3110: "Crédito social CCAF (3110)",
    3181: "Cuota vivienda o educación (3181)",
    3182: "Crédito cooperativas de ahorro (3182)",
    3183: "Otros descuentos autorizados y solicitados por el trabajador (3183)",
    3154: "Cotización adicional trabajo pesado - trabajador (3154)",
    3184: "Donaciones culturales y de reconstrucción (3184)",
    3185: "Otros descuentos (3185)",
    3186: "Pensiones de alimentos (3186)",
    3187: "Descuento mujer casada (3187)",
    3188: "Descuentos por anticipos y préstamos (3188)",
    4151: "AFC - Aporte empleador (4151)",
    4152: "Aporte empleador seguro accidentes del trabajo y Ley SANNA (4152)",
    4131: "Aporte empleador indemnización a todo evento (4131)",
    4154: "Aporte adicional trabajo pesado - empleador (4154)",
    4155: "Aporte empleador seguro invalidez y sobrevivencia (4155)",
    4157: "APVC - Aporte Empleador (4157)",
    5201: "Total haberes (5201)",
    5210: "Total haberes imponibles y tributables (5210)",
    5220: "Total haberes imponibles no tributables (5220)",
    5230: "Total haberes no imponibles y no tributables (5230)",
    5240: "Total haberes no imponibles y tributables (5240)",
    5301: "Total descuentos (5301)",
    5361: "Total descuentos impuestos a las remuneraciones (5361)",
    5362: "Total descuentos impuestos por indemnizaciones (5362)",
    5341: "Total descuentos por cotizaciones del trabajador (5341)",
    5302: "Total otros descuentos (5302)",
    5410: "Total aportes empleador (5410)",
    5501: "Total líquido (5501)",
    5502: "Total indemnizaciones (5502)",
    5564: "Total indemnizaciones tributables (5564)",
    5565: "Total indemnizaciones no tributables (5565)",
}

# ─────────────────────────────────────────────
# FUNCIÓN DE BÚSQUEDA DE COLUMNA POR CÓDIGO
# ─────────────────────────────────────────────
def extraer_codigo(nombre_col):
    """Extrae el código numérico entre paréntesis de un nombre de columna
    (ej: 'Sueldo (2101)' o 'Sueldo(2101)' → 2101). Retorna None si no encuentra."""
    m = re.search(r"\((\d+)\)", str(nombre_col))
    return int(m.group(1)) if m else None

def find_col(df, codigo):
    """Retorna el nombre de columna del DataFrame que contiene (codigo).
    Busca primero por el nombre oficial en LRE_COLUMNAS, luego por patrón (codigo)."""
    nombre_oficial = LRE_COLUMNAS.get(codigo, "")
    if nombre_oficial and nombre_oficial in df.columns:
        return nombre_oficial
    # Fallback: buscar por patrón en caso de variación de nombre
    patron = f"({codigo})"
    for c in df.columns:
        if patron in c:
            return c
    return None

def get_val(row, df, codigo, default=0):
    """Retorna el valor de la columna con ese código para una fila."""
    col = find_col(df, codigo)
    if col and col in row.index:
        v = pd.to_numeric(row[col], errors="coerce")
        return v if pd.notna(v) else default
    return default

def safe_sum_by_codes(df, codigos):
    """Suma columnas identificadas por código numérico."""
    cols = [find_col(df, c) for c in codigos]
    cols = [c for c in cols if c is not None and c in df.columns]
    if not cols:
        return pd.Series(0, index=df.index)
    return df[cols].apply(pd.to_numeric, errors="coerce").fillna(0).sum(axis=1)

CONCEPTOS_AFECTO_AFP = {"mutual", "sis", "trabajoPesaEmpl", "afp", "isapre"}
CONCEPTOS_AFECTO_CES = {"cesAporteCi", "cesAporteSol", "cesEmpleado"}
CONCEPTOS_ID_AFP     = {"sis", "afp", "trabajoPesaEmpl", "cesEmpleado", "cesAporteSol", "cesAporteCi"}


# ─────────────────────────────────────────────
# LECTURA DEL CSV DT (encabezado variable)
# ─────────────────────────────────────────────
def leer_csv_dt(file_obj):
    """Lee el CSV de la DT detectando automáticamente la posición del encabezado."""
    for enc in ("utf-8", "latin-1", "utf-8-sig", "cp1252"):
        try:
            file_obj.seek(0)
            raw = file_obj.read().decode(enc)
            break
        except Exception:
            continue
    else:
        raise ValueError("No se pudo decodificar el archivo CSV.")

    lines = raw.splitlines()
    header_idx = None
    for i, line in enumerate(lines):
        if "Rut trabajador" in line:
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("No se encontró la línea de encabezado en el archivo.")

    data_lines = [lines[header_idx]] + [l for i, l in enumerate(lines) if i != header_idx and l.strip()]
    content = "\n".join(data_lines)
    df = pd.read_csv(StringIO(content), sep=";", dtype=str)
    for col in df.columns:
        converted = pd.to_numeric(df[col], errors="coerce")
        if converted.notna().sum() > df[col].notna().sum() * 0.5:
            df[col] = converted
    return df


def validar_columnas_lre(df):
    """Compara las columnas del CSV contra LRE_COLUMNAS.
    Retorna dos listas: diferencias (nombre distinto) y desconocidas (código no existe)."""
    diferencias = []
    desconocidas = []
    nombres_oficiales = set(LRE_COLUMNAS.values())

    for col in df.columns:
        if col in nombres_oficiales:
            continue  # coincide exactamente
        # Buscar si algún código conocido está en el nombre de la columna
        codigo_encontrado = None
        for cod, nombre_oficial in LRE_COLUMNAS.items():
            if f"({cod})" in col:
                codigo_encontrado = cod
                nombre_oficial_cod = nombre_oficial
                break
        if codigo_encontrado:
            # El código existe pero el nombre difiere
            diferencias.append({
                "Código": codigo_encontrado,
                "Nombre en archivo": col,
                "Nombre oficial LRE": nombre_oficial_cod,
            })
        else:
            # Columna completamente desconocida
            desconocidas.append(col)

    return diferencias, desconocidas


def mostrar_aviso_columnas(diferencias, desconocidas):
    """Muestra avisos de diferencias y columnas desconocidas. 
    Permite agregar columnas nuevas al diccionario en sesión."""

    # ── Aviso 1: nombres distintos (amarillo) ──
    if diferencias:
        st.markdown("""
        <div class="alert-warning">
            ⚠️ <b>Columnas con nombre distinto al LRE oficial.</b>
            El proceso continúa usando el código numérico como referencia.
            Avisa a tu equipo técnico para actualizar el diccionario.
        </div>""", unsafe_allow_html=True)
        with st.expander("📋 Ver detalle de diferencias de nombre"):
            st.dataframe(pd.DataFrame(diferencias), use_container_width=True, hide_index=True)

    # ── Aviso 2: columnas desconocidas (rojo) ──
    if desconocidas:
        st.markdown(f"""
        <div class="alert-error">
            ❌ <b>Se detectaron {len(desconocidas)} columna(s) desconocida(s)</b> que no existen en el diccionario LRE.<br>
            Puedes incorporarlas a continuación para esta sesión. Luego avisa a tu equipo técnico para actualizarlas permanentemente en el código.
        </div>""", unsafe_allow_html=True)

        with st.expander("➕ Incorporar columnas desconocidas al diccionario"):
            for col_desc in desconocidas:
                st.markdown(f"**Columna:** `{col_desc}`")
                col_a, col_b, col_c = st.columns([2, 2, 1])
                with col_a:
                    nombre_ingresado = st.text_input(
                        "Nombre oficial",
                        value=col_desc,
                        key=f"nombre_{col_desc}"
                    )
                with col_b:
                    codigo_ingresado = st.number_input(
                        "Código numérico",
                        min_value=1000,
                        max_value=9999,
                        value=1000,
                        step=1,
                        key=f"codigo_{col_desc}"
                    )
                with col_c:
                    st.markdown("<br>", unsafe_allow_html=True)
                    if st.button("✅ Agregar", key=f"btn_{col_desc}"):
                        LRE_COLUMNAS[int(codigo_ingresado)] = nombre_ingresado
                        st.success(f"Columna `{nombre_ingresado}` agregada con código {codigo_ingresado} para esta sesión.")
                st.markdown("---")



def extraer_fecha_dt(nombre_archivo):
    """
    Intenta extraer yyyy-mm del nombre del archivo.
    Patrones soportados: 'enero_2025', '2025_01', '202501', 'enero-2025', etc.
    Retorna (fecha_str, True) si se encontró, (None, False) si no.
    """
    meses_es = {
        "enero": "01", "febrero": "02", "marzo": "03", "abril": "04",
        "mayo": "05", "junio": "06", "julio": "07", "agosto": "08",
        "septiembre": "09", "octubre": "10", "noviembre": "11", "diciembre": "12"
    }

    nombre = nombre_archivo.lower()

    # Patrón: mes_nombre + año  ej: enero_2025, enero-2025
    for mes_nom, mes_num in meses_es.items():
        patron = rf"{mes_nom}[_\-\s](\d{{4}})"
        m = re.search(patron, nombre)
        if m:
            return f"{m.group(1)}-{mes_num}", True
        # año + mes_nombre  ej: 2025_enero
        patron2 = rf"(\d{{4}})[_\-\s]{mes_nom}"
        m2 = re.search(patron2, nombre)
        if m2:
            return f"{m2.group(1)}-{mes_num}", True

    # Patrón: yyyymm  ej: 202501
    m = re.search(r"(20\d{2})(0[1-9]|1[0-2])", nombre)
    if m:
        return f"{m.group(1)}-{m.group(2)}", True

    # Patrón: yyyy-mm o yyyy_mm
    m = re.search(r"(20\d{2})[_\-](0[1-9]|1[0-2])", nombre)
    if m:
        return f"{m.group(1)}-{m.group(2)}", True

    return None, False


# ─────────────────────────────────────────────
# CARGA DE REFERENCIAS DT
# ─────────────────────────────────────────────
def cargar_referencias_dt(data_dir="data"):
    """Carga todos los archivos de referencia necesarios para el módulo DT."""
    refs = {}
    errores = []
    archivos = {
        "equiv_conceptos":    "equiv_conceptos.xlsx",
        "parametros":         "parametrosMesuales.xlsx",
        "inst_afp":           "inst_afp.xlsx",
        "inst_mutuales":      "inst_mutuales.xlsx",
        "inst_salud":         "inst_salud.xlsx",
        "inst_cajas":         "inst_cajas.xlsx",
        "listado_empresas":   "listado_empresas.xlsx",
    }
    for key, fname in archivos.items():
        path = os.path.join(data_dir, fname)
        if os.path.exists(path):
            refs[key] = pd.read_excel(path)
        else:
            errores.append(fname)
    return refs, errores


# ─────────────────────────────────────────────
# CARGA LISTADO EMPLEADOS (con encabezado en fila 1)
# ─────────────────────────────────────────────
def cargar_empleados(file_obj):
    """
    Lee listado_empleados.xlsx que tiene encabezado real en la fila 1 (índice 0 es título).
    Retorna DataFrame limpio con columnas reales.
    """
    df = pd.read_excel(file_obj, header=1)
    df.columns = [str(c).strip() for c in df.columns]
    return df


def cargar_empresas(file_obj_or_df):
    """
    Lee listado_empresas.xlsx que tiene encabezado real en la fila 1.
    Acepta path, file object o DataFrame ya cargado.
    """
    if isinstance(file_obj_or_df, pd.DataFrame):
        return file_obj_or_df
    df = pd.read_excel(file_obj_or_df, header=1)
    df.columns = [str(c).strip() for c in df.columns]
    return df


# ─────────────────────────────────────────────
# RESOLUCIÓN DE CONTRATOS
# ─────────────────────────────────────────────
def resolver_contrato(df_empleados, rut, fecha_proceso):
    """
    Dado un RUT y la fecha de proceso (yyyy-mm), determina qué contrato aplica.

    Retorna:
        (contrato, empresa_codigo, ok, motivo)
        - ok=True  → se encontró exactamente un contrato válido
        - ok=False → ninguno o más de uno coincide (motivo indica el problema)
    """
    if "Rut" not in df_empleados.columns:
        return "", "", False, "Sin columna Rut"

    emp_rows = df_empleados[df_empleados["Rut"] == rut]
    if emp_rows.empty:
        return "", "", False, "RUT no encontrado en listado de empleados"

    # Si tiene un solo contrato, retornar directamente
    if len(emp_rows) == 1:
        row = emp_rows.iloc[0]
        contrato = row.get("Contrato", "")
        empresa  = str(row.get("Empresa", "")).strip()
        return contrato, empresa, True, ""

    # Múltiples contratos → resolver por intervalo de fechas
    try:
        fp_date = datetime.strptime(fecha_proceso, "%Y-%m")
    except Exception:
        return "", "", False, "Fecha de proceso inválida"

    col_inicio  = next((c for c in ["Fecha Inicio contrato", "Fecha inicio contrato",
                                     "Fecha inicio", "FechaInicio"] if c in emp_rows.columns), None)
    col_termino = next((c for c in ["Fecha término contrato", "Fecha termino contrato",
                                     "Fecha término", "FechaTermino"] if c in emp_rows.columns), None)

    if col_inicio is None:
        return "", "", False, "No se encontró columna de fecha inicio contrato"

    candidatos = []
    for _, cr in emp_rows.iterrows():
        try:
            f_inicio = pd.to_datetime(cr[col_inicio])
            if pd.isna(f_inicio):
                continue

            # Fecha término: si está vacía = contrato indefinido
            f_termino = None
            if col_termino:
                ft = pd.to_datetime(cr[col_termino], errors="coerce")
                if not pd.isna(ft):
                    f_termino = ft

            # Verificar si la fecha de proceso cae en el intervalo
            if f_inicio <= fp_date:
                if f_termino is None or fp_date <= f_termino:
                    candidatos.append(cr)
        except Exception:
            continue

    if len(candidatos) == 1:
        contrato = candidatos[0].get("Contrato", "")
        empresa  = str(candidatos[0].get("Empresa", "")).strip()
        return contrato, empresa, True, ""
    elif len(candidatos) == 0:
        return "", "", False, f"Ningún contrato aplica para el período {fecha_proceso}"
    else:
        return "", "", False, f"Más de un contrato aplica para el período {fecha_proceso} ({len(candidatos)} contratos)"


def construir_log_contratos(df_empleados, ruts_problema, fecha_proceso, motivos):
    """Construye el DataFrame de log para trabajadores con problema de contratos."""
    cols = [c for c in ["Rut", "Nombre", "Empresa", "Contrato",
                         "Fecha Inicio contrato", "Fecha término contrato"]
            if c in df_empleados.columns]
    detalle = df_empleados[df_empleados["Rut"].isin(ruts_problema)][cols].copy()
    detalle["Motivo"] = detalle["Rut"].map(motivos)
    detalle["Período"] = fecha_proceso
    return detalle


# ─────────────────────────────────────────────
# LOOKUP HELPERS
# ─────────────────────────────────────────────
def lookup(df, col_busca, valor, col_trae, default=""):
    """Busca valor en col_busca y retorna col_trae."""
    if df is None or df.empty:
        return default
    try:
        valor_num = pd.to_numeric(valor, errors="coerce")
        mask = df[col_busca] == (valor_num if not pd.isna(valor_num) else valor)
        if mask.any():
            return df.loc[mask.idxmax(), col_trae]
    except Exception:
        pass
    return default


def safe_num(val, default=0):
    """Convierte a número de forma segura."""
    try:
        v = float(val)
        return v if not np.isnan(v) else default
    except Exception:
        return default


# ─────────────────────────────────────────────
# GENERACIÓN DE FILAS DE SALIDA
# ─────────────────────────────────────────────
def generar_filas_dt(df, fecha_proceso, refs, df_empleados, df_empresas_externo=None):
    """
    Genera las filas del archivo de salida desde el CSV de la DT.
    df_empresas_externo: DataFrame de empresas ya cargado correctamente (prioritario).
    """
    filas = []

    equiv        = refs.get("equiv_conceptos", pd.DataFrame())
    params_df    = refs.get("parametros", pd.DataFrame())
    inst_afp     = refs.get("inst_afp", pd.DataFrame())
    inst_mutuales= refs.get("inst_mutuales", pd.DataFrame())
    inst_salud   = refs.get("inst_salud", pd.DataFrame())
    inst_cajas   = refs.get("inst_cajas", pd.DataFrame())

    # Usar df_empresas_externo si se pasó, si no intentar desde refs
    if df_empresas_externo is not None and not df_empresas_externo.empty:
        df_empresas = df_empresas_externo.copy()
    else:
        empresas_raw = refs.get("listado_empresas", pd.DataFrame())
        if isinstance(empresas_raw, pd.DataFrame) and "Empresa" in empresas_raw.columns:
            df_empresas = empresas_raw.copy()
        else:
            df_empresas = cargar_empresas(empresas_raw)

    # Normalizar columnas Nombre y Empresa
    if "Nombre" in df_empresas.columns:
        df_empresas["Nombre"] = df_empresas["Nombre"].astype(str).str.strip()
    if "Empresa" in df_empresas.columns:
        df_empresas["Empresa"] = df_empresas["Empresa"].astype(str).str.strip()

    # ── Parámetros del mes ──
    tope_salud = 0
    tope_afp   = 0
    tope_ces   = 0
    row_params = pd.DataFrame()
    if not params_df.empty and "mes_Proc" in params_df.columns:
        params_df["mes_Proc"] = params_df["mes_Proc"].astype(str).str.strip()
        row_params = params_df[params_df["mes_Proc"] == fecha_proceso]
        if not row_params.empty:
            tope_salud = safe_num(row_params.iloc[0].get("topeSalud_pesos", 0))
            tope_afp   = safe_num(row_params.iloc[0].get("topeImp_pesos_afp", 0))
            tope_ces   = safe_num(row_params.iloc[0].get("topeCes_pesos", 0))

    # ── Mapa de equivalencias: código numérico → concepto_detalle + Tipo ──
    # NOTA: se resuelve por código (código) en vez de nombre exacto de columna,
    # ya que el archivo DT puede traer nombres de columna con variaciones
    # respecto al nombre oficial LRE (espacios, mayúsculas, etc.), igual que
    # find_col(). Comparar por string exacto dejaba cols_conceptos vacío
    # cuando había diferencias de nombre, y por lo tanto no se generaba
    # ninguna fila de salida aunque las validaciones cuadraran.
    equiv_map_by_code = {}  # código (int) → concepto_detalle
    tipo_map = {}           # concepto_detalle → Tipo
    if not equiv.empty and "cod_lre" in equiv.columns and "concepto_detalle" in equiv.columns:
        for _, er in equiv.iterrows():
            codigo_equiv = extraer_codigo(er["cod_lre"])
            concepto     = str(er["concepto_detalle"]).strip()
            tipo         = str(er.get("Tipo", "")).strip()
            # Solo mapear la primera aparición para evitar duplicados de isapre
            if codigo_equiv is not None and codigo_equiv not in equiv_map_by_code:
                equiv_map_by_code[codigo_equiv] = concepto
            tipo_map[concepto] = tipo

    # Resolver cada columna real del df a su concepto, por código numérico
    # (tolerante a diferencias de nombre respecto al oficial LRE)
    equiv_map = {}  # nombre real de columna en df → concepto_detalle
    for col in df.columns:
        codigo_col = extraer_codigo(col)
        if codigo_col is not None and codigo_col in equiv_map_by_code:
            equiv_map[col] = equiv_map_by_code[codigo_col]

    # Columnas del CSV que tienen equivalencia (excluir col salud voluntaria para isapre, se suma manualmente)
    col_salud_vol = find_col(df, COD_SALUD_VOL)
    cols_conceptos = [c for c in df.columns if c in equiv_map and c != col_salud_vol]

    # Conceptos que se incluyen cuando el trabajador tiene licencia mes completo (dias_trab = 0)
    CONCEPTOS_LICENCIA_COMPLETA = {
        "sueldoBase", "gratificacion", "afp", "isapre", "cesEmpleado",
        "impuesto", "totalesEmpl", "mutual", "sis", "cesAporteCi"
    }

    # Días reales del mes (calculado una sola vez)
    try:
        anio_fp, mes_fp = int(fecha_proceso[:4]), int(fecha_proceso[5:7])
        dias_reales_mes = calendar.monthrange(anio_fp, mes_fp)[1]
    except Exception:
        dias_reales_mes = 30

    # Nombres de columna resueltos una sola vez
    col_rut         = find_col(df, COD_RUT)
    col_dias_trab   = find_col(df, COD_DIAS_TRAB)
    col_dias_lic    = find_col(df, COD_DIAS_LIC)
    col_dias_vac    = find_col(df, COD_DIAS_VAC)
    col_sueldo      = find_col(df, COD_SUELDO)
    col_rebaja_zona = find_col(df, COD_REBAJA_ZONA)
    col_salud7      = find_col(df, COD_SALUD7)
    col_afp         = find_col(df, COD_AFP)
    col_ces_trab    = find_col(df, COD_CES_TRAB)
    col_apvi_mod_b  = find_col(df, COD_APVI_MOD_B)
    col_trab_pesado = find_col(df, COD_TRAB_PESADO)
    col_afp_inst    = find_col(df, COD_AFP_INST)
    col_isapre_inst = find_col(df, COD_ISAPRE_INST)
    col_mutual_inst = find_col(df, COD_MUTUAL_INST)
    col_ccaf_inst   = find_col(df, COD_CCAF_INST)

    # ── Registro de problemas de contrato (para log) ──
    ruts_problema = {}   # rut → motivo
    filas = []

    for _, row in df.iterrows():
        rut = str(row.get(col_rut, "")).strip() if col_rut else ""

        # ── Resolver contrato ──
        numero_contrato, empresa_codigo, ok, motivo = resolver_contrato(
            df_empleados, rut, fecha_proceso
        )
        if not ok:
            ruts_problema[rut] = motivo
            continue

        # ── Lookup empresa → busca por Nombre, trae código Empresa ──
        empresa_salida = empresa_codigo
        if empresa_codigo and not df_empresas.empty and "Nombre" in df_empresas.columns:
            empresa_codigo_strip = str(empresa_codigo).strip()
            df_empresas["Nombre"] = df_empresas["Nombre"].astype(str).str.strip()
            emp2 = df_empresas[df_empresas["Nombre"] == empresa_codigo_strip]
            if not emp2.empty:
                empresa_salida = str(emp2.iloc[0].get("Empresa", empresa_codigo)).strip()

        # ── Valores base del trabajador ──
        dias_trab   = safe_num(row.get(col_dias_trab, 0) if col_dias_trab else 0)
        dias_lic    = safe_num(row.get(col_dias_lic, 0) if col_dias_lic else 0)
        dias_vac    = safe_num(row.get(col_dias_vac, 0) if col_dias_vac else 0)
        sueldo      = safe_num(row.get(col_sueldo, 0) if col_sueldo else 0)
        rebaja_zona = safe_num(row.get(col_rebaja_zona, 0) if col_rebaja_zona else 0)

        monto_init = round((sueldo / dias_trab) * 30, 0) if dias_trab > 0 else 0

        # ── Suma haberes afectos (para Afecto) ──
        conceptos_haber_afecto = [c for c in cols_conceptos
                                   if tipo_map.get(equiv_map.get(c, ""), "") == "Haber afecto"]
        suma_haber_afecto = sum(safe_num(row.get(c, 0)) for c in conceptos_haber_afecto)

        # ── Suma haberes exentos (para Rentas no gravadas) ──
        conceptos_haber_exento = [c for c in cols_conceptos
                                   if tipo_map.get(equiv_map.get(c, ""), "") == "Haber exento"]
        suma_haber_exento = sum(safe_num(row.get(c, 0)) for c in conceptos_haber_exento)

        # ── Monto isapre (caso especial: suma dos columnas) ──
        monto_isapre = (safe_num(row.get(col_salud7, 0) if col_salud7 else 0) +
                        safe_num(row.get(col_salud_vol, 0) if col_salud_vol else 0))

        # ── Total rebajas por LLSS (solo para concepto impuesto) ──
        salud_trab = (safe_num(row.get(col_salud7, 0) if col_salud7 else 0) +
                      safe_num(row.get(col_salud_vol, 0) if col_salud_vol else 0))
        salud_tope = min(salud_trab, tope_salud) if tope_salud > 0 else salud_trab
        total_rebajas_llss = (
            safe_num(row.get(col_afp, 0) if col_afp else 0)
            + safe_num(row.get(col_ces_trab, 0) if col_ces_trab else 0)
            + safe_num(row.get(col_apvi_mod_b, 0) if col_apvi_mod_b else 0)
            + safe_num(row.get(col_trab_pesado, 0) if col_trab_pesado else 0)
            + salud_tope
        )

        # ── Código de institución del trabajador ──
        cod_afp_inst    = safe_num(row.get(col_afp_inst, 0) if col_afp_inst else 0)
        cod_isapre_inst = safe_num(row.get(col_isapre_inst, 0) if col_isapre_inst else 0)
        cod_mutual_inst = safe_num(row.get(col_mutual_inst, 0) if col_mutual_inst else 0)
        cod_ccaf_inst   = safe_num(row.get(col_ccaf_inst, 0) if col_ccaf_inst else 0)

        # ── Lookup instituciones ──
        id_afp_trab    = lookup(inst_afp,      "cod_lre", cod_afp_inst,    "id_afp",    "")
        id_isapre_trab = lookup(inst_salud,    "cod_lre", cod_isapre_inst, "id_inst",   "")
        id_mutual_trab = lookup(inst_mutuales, "cod_lre", cod_mutual_inst, "id_mutual", "")
        id_ccaf_trab   = lookup(inst_cajas,    "cod_lre", cod_ccaf_inst,   "id_ccaf",   "")
        cot_afp_trab   = lookup(inst_afp,      "cod_lre", cod_afp_inst,    "cot_afp",   0)

        # ── Cotización mutual (triangulación empleado → empresa) ──
        cot_mutual = 0.93
        if empresa_salida and not df_empresas.empty and "Empresa" in df_empresas.columns:
            emp2 = df_empresas[df_empresas["Empresa"] == empresa_salida]
            if not emp2.empty and "Cotización Mutual" in df_empresas.columns:
                cot_mutual = safe_num(emp2.iloc[0].get("Cotización Mutual", 0.93), 0.93)

        licencia_mes_completo = dias_trab == 0

        # ── Agrupar montos por concepto (suma columnas que mapean al mismo concepto) ──
        montos_por_concepto = {}
        for col_csv in cols_conceptos:
            id_concepto = equiv_map.get(col_csv, "")
            if not id_concepto:
                continue
            if id_concepto == "isapre":
                monto = monto_isapre
            else:
                monto = safe_num(row.get(col_csv, 0))
            montos_por_concepto[id_concepto] = montos_por_concepto.get(id_concepto, 0) + monto

        # ── Generar fila por cada concepto ──
        conceptos_siempre = {"impuesto", "cesEmpleado"}
        if licencia_mes_completo:
            conceptos_siempre = conceptos_siempre | CONCEPTOS_LICENCIA_COMPLETA
            # Asegurar que todos los conceptos de licencia completa estén presentes
            # aunque su columna no exista en el CSV o su monto sea 0
            for c in CONCEPTOS_LICENCIA_COMPLETA:
                if c not in montos_por_concepto:
                    montos_por_concepto[c] = 0

        for id_concepto, monto in montos_por_concepto.items():

            # Si licencia mes completo, solo incluir conceptos permitidos
            if licencia_mes_completo and id_concepto not in CONCEPTOS_LICENCIA_COMPLETA:
                continue

            # Saltar si monto es 0, excepto conceptos que siempre se incluyen
            if monto == 0 and id_concepto not in conceptos_siempre:
                continue

            # ── Id de institución ──
            id_institucion = ""
            if id_concepto in CONCEPTOS_ID_AFP:
                id_institucion = id_afp_trab
            elif id_concepto == "isapre":
                id_institucion = id_isapre_trab
            elif id_concepto == "mutual":
                id_institucion = id_mutual_trab
            elif id_concepto == "cajaCred":
                id_institucion = id_ccaf_trab

            # ── Afecto ──
            if id_concepto in CONCEPTOS_AFECTO_AFP:
                afecto = min(suma_haber_afecto, tope_afp) if tope_afp > 0 else suma_haber_afecto
            elif id_concepto in CONCEPTOS_AFECTO_CES:
                afecto = min(suma_haber_afecto, tope_ces) if tope_ces > 0 else suma_haber_afecto
            elif id_concepto == "totalesEmpl":
                afecto = suma_haber_afecto
            elif id_concepto == "impuesto":
                afecto = suma_haber_afecto - total_rebajas_llss
            else:
                afecto = 0

            # ── Cotización de jubilación ──
            cot_jubilacion = ""
            if id_concepto == "cesEmpleado":
                cot_jubilacion = 0.6
            elif id_concepto == "afp":
                cot_jubilacion = cot_afp_trab
            elif id_concepto == "isapre" and id_institucion == "fonasa":
                cot_jubilacion = monto
            elif id_concepto == "mutual":
                cot_jubilacion = cot_mutual
            elif id_concepto == "sis":
                cot_jubilacion = safe_num(row_params.iloc[0].get("sis", 0)) if not row_params.empty else 0
            elif id_concepto == "totalesEmpl":
                cot_jubilacion = min(suma_haber_afecto, tope_afp) if tope_afp > 0 else suma_haber_afecto

            # ── Rentas no gravadas (solo si concepto = impuesto) ──
            rentas_no_grav = suma_haber_exento if id_concepto == "impuesto" else 0

            # ── Total rebajas LLSS (solo si concepto = impuesto) ──
            rebajas_llss = total_rebajas_llss if id_concepto == "impuesto" else 0

            filas.append({
                "Fecha de proceso":        fecha_proceso,
                "Id empleado":             rut,
                "Número de contrato":      numero_contrato,
                "Id del concepto":         id_concepto,
                "Monto del concepto":      monto,
                "Afecto":                  afecto,
                "Id de institución":       id_institucion,
                "Cotización de jubilación": cot_jubilacion,
                "Días de licencias":       dias_lic,
                "Días trabajados":         dias_trab,
                "Fecha de aplicación":     fecha_proceso,
                "Empresa":                 empresa_salida,
                "Total de rebajas por LLSS": rebajas_llss,
                "Rentas no gravadas":      rentas_no_grav,
                "Rebaja por zona extrema": rebaja_zona,
                "Jornada":                 "C",
                "Días de vacaciones":      dias_vac,
                "Monto Init":              monto_init,
                "Fase":                    1,
            })

        # ── Fila adicional: licenciaDias (si días de licencia > 0) ──
        if dias_lic > 0:
            filas.append({
                "Fecha de proceso":          fecha_proceso,
                "Id empleado":               rut,
                "Número de contrato":        numero_contrato,
                "Id del concepto":           "licenciaDias",
                "Monto del concepto":        dias_lic,
                "Afecto":                    0,
                "Id de institución":         "",
                "Cotización de jubilación":  "",
                "Días de licencias":         dias_lic,
                "Días trabajados":           dias_trab,
                "Fecha de aplicación":       fecha_proceso,
                "Empresa":                   empresa_salida,
                "Total de rebajas por LLSS": 0,
                "Rentas no gravadas":        0,
                "Rebaja por zona extrema":   rebaja_zona,
                "Jornada":                   "C",
                "Días de vacaciones":        dias_vac,
                "Monto Init":                monto_init,
                "Fase":                      1,
            })
    df_log_contratos = pd.DataFrame()
    if ruts_problema:
        df_log_contratos = construir_log_contratos(
            df_empleados, list(ruts_problema.keys()), fecha_proceso, ruts_problema
        )

    return pd.DataFrame(filas), df_log_contratos


# ─────────────────────────────────────────────
# GENERACIÓN EXCEL DE SALIDA
# ─────────────────────────────────────────────
def generar_excel_dt(df_salida):
    """Genera el Excel de salida con formato Rex+."""
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Liquidaciones"

    header_fill = PatternFill("solid", fgColor="1A2744")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(
        bottom=Side(style="thin", color="E8EDF5"),
        right=Side(style="thin", color="E8EDF5")
    )

    cols = list(df_salida.columns)
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = max(len(col) + 4, 14)

    for ri, row in enumerate(df_salida.itertuples(index=False), 2):
        fill = PatternFill("solid", fgColor="EAF0F8") if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    wb.save(output)
    return output.getvalue()


# ─────────────────────────────────────────────
# GENERACIÓN EXCEL LOG MÚLTIPLES CONTRATOS
# ─────────────────────────────────────────────
def generar_excel_log(df_log):
    """Genera el Excel de log de trabajadores con múltiples contratos."""
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Múltiples contratos"

    header_fill = PatternFill("solid", fgColor="C53030")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    cols = list(df_log.columns)
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = max(len(col) + 4, 18)

    for ri, row in enumerate(df_log.itertuples(index=False), 2):
        fill = PatternFill("solid", fgColor="FFF5F5") if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    wb.save(output)
    return output.getvalue()




def validar_cuadraturas_dt(df, nombre_archivo):
    """Ejecuta las 13 validaciones del instructivo DT."""
    errores = []
    tol = 1

    df = df.copy()

    # Excluir trabajadores con licencia mes completo (dias_trabajados = 0)
    col_dias_trab = find_col(df, COD_DIAS_TRAB)
    if col_dias_trab and col_dias_trab in df.columns:
        df = df[pd.to_numeric(df[col_dias_trab], errors="coerce").fillna(0) != 0]

    col_rut = find_col(df, COD_RUT) or ""

    def sv(df, codigo):
        """Suma una columna por código, retorna Serie."""
        c = find_col(df, codigo)
        if c and c in df.columns:
            return pd.to_numeric(df[c], errors="coerce").fillna(0)
        return pd.Series(0, index=df.index)

    def detalle_cols_dt(row, codigos):
        """Retorna string con nombre_col: monto para columnas con valor != 0."""
        partes = []
        for cod in codigos:
            c = find_col(df, cod)
            if c and c in row.index:
                v = pd.to_numeric(row[c], errors="coerce")
                if pd.notna(v) and v != 0:
                    partes.append(f"{c}: {round(v, 2)}")
        return " | ".join(partes) if partes else "(sin valores)"

    def registrar(df, mask, calc, cod_ctrl, codigo_val, descripcion, codes_detalle, col_rut):
        """Registra errores para las filas que no cumplen la validación."""
        col_ctrl = find_col(df, cod_ctrl)
        if not col_ctrl or col_ctrl not in df.columns:
            return []
        ctrl = pd.to_numeric(df[col_ctrl], errors="coerce").fillna(0)
        filas = []
        for _, row in df[mask].iterrows():
            filas.append({
                "Archivo":          nombre_archivo,
                "RUT":              row.get(col_rut, "N/D"),
                "Validación":       codigo_val,
                "Descripción":      descripcion,
                "Columnas sumadas": detalle_cols_dt(row, codes_detalle),
                "Valor calculado":  round(calc[row.name], 2),
                "Columna control":  col_ctrl,
                "Valor control":    round(ctrl[row.name], 2),
                "Diferencia":       round(calc[row.name] - ctrl[row.name], 2),
            })
        return filas

    # ── Cálculos intermedios ──
    # Totales de secciones
    calc_5220 = safe_sum_by_codes(df, [2201, 2202, 2203, 2204])
    calc_5230 = safe_sum_by_codes(df, [2301, 2302, 2303, 2304, 2305, 2311, 2306,
                                        2307, 2308, 2309, 2347, 2310, 2312, 2313,
                                        2314, 2315, 2316, 2331])
    calc_5301 = safe_sum_by_codes(df, [3141, 3143, 3144, 3151, 3146, 3147, 3155,
                                        3156, 3157, 3158, 3161, 3162, 3163, 3165,
                                        3166, 3167, 3171, 3172, 3173, 3174, 3175,
                                        3176, 3177, 3178, 3179, 3180, 3110, 3181,
                                        3182, 3183, 3154, 3184, 3185, 3186, 3187, 3188])
    calc_5341 = safe_sum_by_codes(df, [3141, 3143, 3144, 3146, 3151, 3154, 3155, 3156, 3157, 3158])
    calc_5361 = safe_sum_by_codes(df, [3161, 3165])
    calc_5362 = sv(df, 3162)
    calc_5302 = calc_5301 - calc_5361 - calc_5362 - calc_5341
    calc_5410 = safe_sum_by_codes(df, [4151, 4152, 4131, 4154, 4155, 4157])
    calc_5502 = safe_sum_by_codes(df, [2313, 2314, 2315, 2316, 2331, 2417, 2418])
    calc_5564 = safe_sum_by_codes(df, [2417, 2418])
    calc_5565 = safe_sum_by_codes(df, [2313, 2314, 2315, 2316, 2331])
    calc_5201 = sv(df, 5210) + calc_5220 + calc_5230 + sv(df, 5240)
    calc_5501 = calc_5201 - calc_5301

    # ── Validaciones ──
    validaciones = [
        # (codigo_val, calc, cod_ctrl, descripcion, codes_detalle)
        ("V1",  calc_5201, 5201, "5210+5220+5230+5240 ≠ Total haberes (5201)",
         [5210, 5220, 5230, 5240]),
        ("V2",  calc_5220, 5220, "2201+2202+2203+2204 ≠ Total haberes imponibles no tributables (5220)",
         [2201, 2202, 2203, 2204]),
        ("V3",  calc_5230, 5230, "Suma sección 23xx ≠ Total haberes no imponibles y no tributables (5230)",
         [2301, 2302, 2303, 2304, 2305, 2311, 2306, 2307, 2308, 2309, 2347, 2310, 2312, 2313, 2314, 2315, 2316, 2331]),
        ("V4",  calc_5301, 5301, "Suma descuentos ≠ Total descuentos (5301)",
         [3141, 3143, 3144, 3151, 3146, 3147, 3155, 3156, 3157, 3158, 3161, 3162, 3163,
          3165, 3166, 3167, 3171, 3172, 3173, 3174, 3175, 3176, 3177, 3178, 3179, 3180,
          3110, 3181, 3182, 3183, 3154, 3184, 3185, 3186, 3187, 3188]),
        ("V5",  calc_5361, 5361, "3161+3165 ≠ Total descuentos impuestos remuneraciones (5361)",
         [3161, 3165]),
        ("V6",  calc_5362, 5362, "3162 ≠ Total descuentos impuestos indemnizaciones (5362)",
         [3162]),
        ("V7",  calc_5341, 5341, "Suma cotizaciones ≠ Total descuentos cotizaciones trabajador (5341)",
         [3141, 3143, 3144, 3146, 3151, 3154, 3155, 3156, 3157, 3158]),
        ("V8",  calc_5302, 5302, "5301-5361-5362-5341 ≠ Total otros descuentos (5302)",
         []),
        ("V9",  calc_5410, 5410, "Suma aportes empleador ≠ Total aportes empleador (5410)",
         [4151, 4152, 4131, 4154, 4155, 4157]),
        ("V10", calc_5501, 5501, "5201-5301 ≠ Total líquido (5501)",
         []),
        ("V11", calc_5502, 5502, "Suma indemnizaciones ≠ Total indemnizaciones (5502)",
         [2313, 2314, 2315, 2316, 2331, 2417, 2418]),
        ("V12", calc_5564, 5564, "2417+2418 ≠ Total indemnizaciones tributables (5564)",
         [2417, 2418]),
        ("V13", calc_5565, 5565, "Suma ≠ Total indemnizaciones no tributables (5565)",
         [2313, 2314, 2315, 2316, 2331]),
    ]

    for codigo_val, calc, cod_ctrl, descripcion, codes_detalle in validaciones:
        col_ctrl = find_col(df, cod_ctrl)
        if not col_ctrl or col_ctrl not in df.columns:
            continue
        ctrl = pd.to_numeric(df[col_ctrl], errors="coerce").fillna(0)
        mask = (calc - ctrl).abs() > tol
        for _, row in df[mask].iterrows():
            errores.append({
                "Archivo":          nombre_archivo,
                "RUT":              row.get(col_rut, "N/D"),
                "Validación":       codigo_val,
                "Descripción":      descripcion,
                "Columnas sumadas": detalle_cols_dt(row, codes_detalle) if codes_detalle else "Ver descripción",
                "Valor calculado":  round(calc[row.name], 2),
                "Columna control":  col_ctrl,
                "Valor control":    round(ctrl[row.name], 2),
                "Diferencia":       round(calc[row.name] - ctrl[row.name], 2),
            })

    return errores


# ─────────────────────────────────────────────
# INTERFAZ STREAMLIT DEL MÓDULO DT
# ─────────────────────────────────────────────
def render_modulo_dt(refs_compartidas):
    """
    Renderiza la interfaz del módulo DT dentro del tab de Streamlit.
    refs_compartidas: dict con archivos de referencia ya cargados por app_migracion.
    """
    st.markdown('<div class="section-title">🏛️ Migración Declaración Jurada DT</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-sub">Procesa el archivo descargado directamente desde la Dirección del Trabajo y genera el archivo de liquidaciones en detalle Rex+.</div>', unsafe_allow_html=True)

    # Pasos
    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown("""
        <div class="step-card">
            <div class="step-label">PASO 1</div>
            <div class="step-title">Subir archivo DT</div>
            <div class="step-desc">Archivo CSV descargado desde el portal de la Dirección del Trabajo.</div>
        </div>""", unsafe_allow_html=True)
    with col2:
        st.markdown("""
        <div class="step-card">
            <div class="step-label">PASO 2</div>
            <div class="step-title">Subir listado de empleados</div>
            <div class="step-desc">Listado de empleados del período a procesar exportado desde Rex+.</div>
        </div>""", unsafe_allow_html=True)
    with col3:
        st.markdown("""
        <div class="step-card">
            <div class="step-label">PASO 3</div>
            <div class="step-title">Descargar resultado</div>
            <div class="step-desc">Se genera el archivo Excel de liquidaciones listo para importar.</div>
        </div>""", unsafe_allow_html=True)

    st.markdown('<hr class="rex-divider">', unsafe_allow_html=True)

    # ── Uploaders en 2x2 ──
    col_up1, col_up2 = st.columns(2)

    with col_up1:
        st.markdown("#### 📤 Archivo CSV de la Dirección del Trabajo")
        archivo_dt = st.file_uploader(
            "Selecciona el archivo CSV de la DT",
            type=["csv"],
            accept_multiple_files=False,
            key="dt_csv_upload",
            help="Archivo descargado desde el portal de la Dirección del Trabajo."
        )
        if archivo_dt:
            st.markdown(f'<div class="alert-success">✅ <b>{archivo_dt.name}</b></div>', unsafe_allow_html=True)

    with col_up2:
        st.markdown("#### 👥 Listado de empleados del período")
        archivo_empleados = st.file_uploader(
            "Sube el archivo listado_empleados.xlsx del período",
            type=["xlsx"],
            accept_multiple_files=False,
            key="dt_empleados_upload",
            help="Exportado desde Rex+. Debe contener columnas: Rut, Nombre, Empresa, Contrato."
        )
        if archivo_empleados:
            st.markdown(f'<div class="alert-success">✅ <b>{archivo_empleados.name}</b></div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="alert-warning">⚠️ Requerido para ejecutar el proceso.</div>', unsafe_allow_html=True)

    col_up3, col_up4 = st.columns(2)

    with col_up3:
        st.markdown("#### 🏢 Listado de empresas del período")
        archivo_empresas = st.file_uploader(
            "Sube el archivo listado_empresas.xlsx del período",
            type=["xlsx"],
            accept_multiple_files=False,
            key="dt_empresas_upload",
            help="Exportado desde Rex+. Debe contener columnas: Empresa, Nombre, Cotización Mutual."
        )
        if archivo_empresas:
            st.markdown(f'<div class="alert-success">✅ <b>{archivo_empresas.name}</b></div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="alert-warning">⚠️ Requerido para ejecutar el proceso.</div>', unsafe_allow_html=True)

    with col_up4:
        st.markdown("#### 📅 Parámetros mensuales")
        archivo_params_dt = st.file_uploader(
            "Sube el archivo parametrosMesuales.xlsx del período",
            type=["xlsx"],
            accept_multiple_files=False,
            key="dt_params_upload",
            help="Archivo con los parámetros legales del mes a procesar."
        )
        if archivo_params_dt:
            st.markdown(f'<div class="alert-success">✅ <b>{archivo_params_dt.name}</b></div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="alert-warning">⚠️ Requerido para ejecutar el proceso.</div>', unsafe_allow_html=True)

    if not archivo_dt or not archivo_empleados or not archivo_empresas or not archivo_params_dt:
        return

    st.markdown(f'<div class="alert-success">✅ Archivo DT cargado: <b>{archivo_dt.name}</b></div>', unsafe_allow_html=True)

    # ── Determinar fecha de proceso ──
    fecha_proceso, fecha_ok = extraer_fecha_dt(archivo_dt.name)

    if not fecha_ok:
        st.markdown("""
        <div class="alert-warning">
            ⚠️ <b>No se pudo determinar la fecha de proceso</b> desde el nombre del archivo.<br>
            Por favor selecciona el mes y año correspondiente.
        </div>""", unsafe_allow_html=True)

        col_m, col_a, _ = st.columns([1, 1, 3])
        with col_m:
            mes_sel = st.selectbox("Mes", options=list(range(1, 13)),
                                   format_func=lambda x: f"{x:02d} - {['Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre'][x-1]}",
                                   key="dt_mes_sel")
        with col_a:
            anio_actual = datetime.now().year
            anio_sel = st.selectbox("Año", options=list(range(anio_actual - 5, anio_actual + 2)),
                                    index=5, key="dt_anio_sel")
        fecha_proceso = f"{anio_sel}-{mes_sel:02d}"

    st.markdown(f'<div class="alert-success">📅 Fecha de proceso: <b>{fecha_proceso}</b></div>', unsafe_allow_html=True)

    # ── Botón ejecutar ──
    if st.button("▶ Ejecutar proceso DT", key="dt_btn_ejecutar"):
        with st.spinner("Procesando archivo..."):
            try:
                # Leer CSV DT
                df_dt = leer_csv_dt(archivo_dt)

                # Validar columnas contra LRE_COLUMNAS
                diferencias_cols, desconocidas_cols = validar_columnas_lre(df_dt)
                if diferencias_cols or desconocidas_cols:
                    mostrar_aviso_columnas(diferencias_cols, desconocidas_cols)

                # Leer empleados
                df_empleados = cargar_empleados(archivo_empleados)

                # Leer empresas
                df_empresas_periodo = pd.read_excel(archivo_empresas, header=1)
                df_empresas_periodo.columns = [str(c).strip() for c in df_empresas_periodo.columns]
                if "Nombre" in df_empresas_periodo.columns:
                    df_empresas_periodo["Nombre"] = df_empresas_periodo["Nombre"].astype(str).str.strip()
                if "Empresa" in df_empresas_periodo.columns:
                    df_empresas_periodo["Empresa"] = df_empresas_periodo["Empresa"].astype(str).str.strip()

                # Leer parámetros mensuales
                df_params_periodo = pd.read_excel(archivo_params_dt)

                # Verificar que el mes existe en los parámetros
                if "mes_Proc" in df_params_periodo.columns:
                    df_params_periodo["mes_Proc"] = df_params_periodo["mes_Proc"].astype(str).str.strip()
                    if fecha_proceso not in df_params_periodo["mes_Proc"].values:
                        st.markdown(f'<div class="alert-error">❌ <b>No se encontraron parámetros para {fecha_proceso}</b> en el archivo subido.</div>', unsafe_allow_html=True)
                        st.stop()

                # Inyectar en refs
                refs_dt = dict(refs_compartidas)
                refs_dt["listado_empresas"] = df_empresas_periodo
                refs_dt["parametros"] = df_params_periodo

                # Ejecutar validaciones de cuadratura
                errores_val = validar_cuadraturas_dt(df_dt, archivo_dt.name)

            except Exception as e:
                st.markdown(f'<div class="alert-error">❌ Error al leer los archivos: <b>{e}</b></div>', unsafe_allow_html=True)
                return

        st.markdown('<hr class="rex-divider">', unsafe_allow_html=True)
        st.markdown("### 🔍 Resultado de validaciones")

        # ── Mostrar errores de validación si hay ──
        if errores_val:
            st.markdown(f"""
            <div class="alert-error">
                ❌ <b>No se puede generar el archivo de salida.</b><br>
                Se encontraron <b>{len(errores_val)} error(es)</b> de validación en los registros procesados.
            </div>""", unsafe_allow_html=True)

            with st.expander("📋 Ver log de errores detallado"):
                df_errores = pd.DataFrame(errores_val)
                st.dataframe(df_errores, use_container_width=True, hide_index=True)
                excel_log = io.BytesIO()
                df_errores.to_excel(excel_log, index=False, sheet_name="Log errores")
                excel_log.seek(0)
                st.download_button(
                    label="⬇️ Descargar log de errores (.xlsx)",
                    data=excel_log,
                    file_name=f"log_errores_dt_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dt_btn_log_errores"
                )
            return

        st.markdown("""
        <div class="alert-success">
            ✅ <b>Todas las validaciones se cumplieron correctamente.</b><br>
            Los registros cuadran sin diferencias.
        </div>""", unsafe_allow_html=True)

        # ── Generar archivo de salida ──
        with st.spinner("Generando archivo de salida..."):
            try:
                df_salida, df_log_contratos = generar_filas_dt(df_dt, fecha_proceso, refs_dt, df_empleados, df_empresas_externo=df_empresas_periodo)
            except Exception as e:
                st.markdown(f'<div class="alert-error">❌ Error al generar el archivo de salida: <b>{e}</b></div>', unsafe_allow_html=True)
                return

        # ── Mostrar log de problemas de contrato si hay ──
        if not df_log_contratos.empty:
            st.markdown(f"""
            <div class="alert-warning">
                ⚠️ <b>{len(df_log_contratos["Rut"].unique())} trabajador(es)</b> fueron excluidos por problemas en la resolución de contrato.<br>
                Descarga el log para revisarlos.
            </div>""", unsafe_allow_html=True)

            with st.expander(f"👁️ Ver trabajadores excluidos por contrato ({len(df_log_contratos['Rut'].unique())})"):
                st.dataframe(df_log_contratos, use_container_width=True, hide_index=True)

            log_cont_bytes = generar_excel_log(df_log_contratos)
            st.download_button(
                label="⬇️ Descargar log_multiples_contratos.xlsx",
                data=log_cont_bytes,
                file_name=f"log_multiples_contratos_{fecha_proceso}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dt_btn_log_contratos"
            )

        if df_salida.empty:
            st.markdown('<div class="alert-warning">⚠️ No se generaron registros. Verifica que el archivo y los parámetros sean correctos.</div>', unsafe_allow_html=True)
            return

        st.markdown(f"""
        <div class="alert-success">
            ✅ <b>Archivo generado exitosamente.</b><br>
            Se procesaron <b>{df_salida["Id empleado"].nunique()} trabajadores</b> y se generaron <b>{len(df_salida)} registros</b>.
        </div>""", unsafe_allow_html=True)

        with st.expander("👁️ Vista previa del archivo de salida"):
            st.dataframe(df_salida.head(50), use_container_width=True, hide_index=True)

        excel_bytes = generar_excel_dt(df_salida)
        st.download_button(
            label="⬇️ Descargar archivo de salida (.xlsx)",
            data=excel_bytes,
            file_name=f"liquidaciones_dt_{fecha_proceso}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dt_btn_descarga"
        )
