"""
modulo_dt.py
Procesamiento del archivo de la Dirección del Trabajo (DT)
para generación del archivo de liquidaciones en detalle Rex+.
"""

import re
import io
import os
import pandas as pd
import numpy as np
import streamlit as st
from io import StringIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


# ─────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────
COL_RUT           = "Rut trabajador(1101)"
COL_DIAS_TRAB     = "Nro días trabajados en el mes(1115)"
COL_DIAS_LIC      = "Nro días de licencia médica en el mes(1116)"
COL_DIAS_VAC      = "Nro días de vacaciones en el mes(1117)"
COL_SUELDO        = "Sueldo(2101)"
COL_SALUD7        = "Cotización obligatoria salud 7%(3143)"
COL_SALUD_VOL     = "Cotización voluntaria para salud(3144)"
COL_AFP           = "Cotización obligatoria previsional (AFP o IPS)(3141)"
COL_CES_TRAB      = "Cotización AFC - trabajador(3151)"
COL_APVI_MOD_B    = "Cotización APVi Mod B hasta UF50(3156)"
COL_TRAB_PESADO   = "Cotización adicional trabajo pesado - trabajador(3154)"
COL_REBAJA_ZONA   = "Rebaja zona extrema DL 889 (3167)"
COL_AFP_INST      = "AFP(1141)"
COL_ISAPRE_INST   = "FONASA - ISAPRE(1143)"
COL_MUTUAL_INST   = "Org. administrador ley 16.744(1152)"
COL_CCAF_INST     = "CCAF(1110)"

CONCEPTOS_AFECTO_AFP = {"mutual", "sis", "trabajoPesaEmpl", "afp", "isapre"}
CONCEPTOS_AFECTO_CES = {"cesAporteCi", "cesAporteSol", "cesEmpleado"}
CONCEPTOS_ID_AFP     = {"sis", "afp", "trabajoPesaEmpl", "cesEmpleado", "cesAporteSol", "cesAporteCi"}


# ─────────────────────────────────────────────
# LECTURA DEL CSV DT (encabezado variable)
# ─────────────────────────────────────────────
def leer_csv_dt(file_obj):
    """Lee el CSV de la DT detectando automáticamente la posición del encabezado."""
    for enc in ("utf-8", "latin-1", "utf-8-sig", "cp1252"):
        try:
            file_obj.seek(0)
            raw = file_obj.read().decode(enc)
            break
        except Exception:
            continue
    else:
        raise ValueError("No se pudo decodificar el archivo CSV.")

    lines = raw.splitlines()
    header_idx = None
    for i, line in enumerate(lines):
        if "Rut trabajador" in line:
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("No se encontró la línea de encabezado en el archivo.")

    data_lines = [lines[header_idx]] + [l for i, l in enumerate(lines) if i != header_idx and l.strip()]
    content = "\n".join(data_lines)
    df = pd.read_csv(StringIO(content), sep=";", dtype=str)
    for col in df.columns:
        converted = pd.to_numeric(df[col], errors="coerce")
        if converted.notna().sum() > df[col].notna().sum() * 0.5:
            df[col] = converted
    return df


# ─────────────────────────────────────────────
# EXTRACCIÓN DE FECHA DESDE NOMBRE DE ARCHIVO
# ─────────────────────────────────────────────
def extraer_fecha_dt(nombre_archivo):
    """
    Intenta extraer yyyy-mm del nombre del archivo.
    Patrones soportados: 'enero_2025', '2025_01', '202501', 'enero-2025', etc.
    Retorna (fecha_str, True) si se encontró, (None, False) si no.
    """
    meses_es = {
        "enero": "01", "febrero": "02", "marzo": "03", "abril": "04",
        "mayo": "05", "junio": "06", "julio": "07", "agosto": "08",
        "septiembre": "09", "octubre": "10", "noviembre": "11", "diciembre": "12"
    }

    nombre = nombre_archivo.lower()

    # Patrón: mes_nombre + año  ej: enero_2025, enero-2025
    for mes_nom, mes_num in meses_es.items():
        patron = rf"{mes_nom}[_\-\s](\d{{4}})"
        m = re.search(patron, nombre)
        if m:
            return f"{m.group(1)}-{mes_num}", True
        # año + mes_nombre  ej: 2025_enero
        patron2 = rf"(\d{{4}})[_\-\s]{mes_nom}"
        m2 = re.search(patron2, nombre)
        if m2:
            return f"{m2.group(1)}-{mes_num}", True

    # Patrón: yyyymm  ej: 202501
    m = re.search(r"(20\d{2})(0[1-9]|1[0-2])", nombre)
    if m:
        return f"{m.group(1)}-{m.group(2)}", True

    # Patrón: yyyy-mm o yyyy_mm
    m = re.search(r"(20\d{2})[_\-](0[1-9]|1[0-2])", nombre)
    if m:
        return f"{m.group(1)}-{m.group(2)}", True

    return None, False


# ─────────────────────────────────────────────
# CARGA DE REFERENCIAS DT
# ─────────────────────────────────────────────
def cargar_referencias_dt(data_dir="data"):
    """Carga todos los archivos de referencia necesarios para el módulo DT."""
    refs = {}
    errores = []
    archivos = {
        "equiv_conceptos":    "equiv_conceptos.xlsx",
        "parametros":         "parametrosMesuales.xlsx",
        "inst_afp":           "inst_afp.xlsx",
        "inst_mutuales":      "inst_mutuales.xlsx",
        "inst_salud":         "inst_salud.xlsx",
        "inst_cajas":         "inst_cajas.xlsx",
        "listado_empresas":   "listado_empresas.xlsx",
    }
    for key, fname in archivos.items():
        path = os.path.join(data_dir, fname)
        if os.path.exists(path):
            refs[key] = pd.read_excel(path)
        else:
            errores.append(fname)
    return refs, errores


# ─────────────────────────────────────────────
# CARGA LISTADO EMPLEADOS (con encabezado en fila 1)
# ─────────────────────────────────────────────
def cargar_empleados(file_obj):
    """
    Lee listado_empleados.xlsx que tiene encabezado real en la fila 1 (índice 0 es título).
    Retorna DataFrame limpio con columnas reales.
    """
    df = pd.read_excel(file_obj, header=1)
    df.columns = [str(c).strip() for c in df.columns]
    return df


def cargar_empresas(file_obj_or_df):
    """
    Lee listado_empresas.xlsx que tiene encabezado real en la fila 1.
    Acepta path, file object o DataFrame ya cargado.
    """
    if isinstance(file_obj_or_df, pd.DataFrame):
        return file_obj_or_df
    df = pd.read_excel(file_obj_or_df, header=1)
    df.columns = [str(c).strip() for c in df.columns]
    return df


# ─────────────────────────────────────────────
# DETECCIÓN DE MÚLTIPLES CONTRATOS
# ─────────────────────────────────────────────
def detectar_multiples_contratos(df_empleados):
    """
    Detecta RUTs con más de un contrato en el listado de empleados.
    Retorna (set de ruts_ok, DataFrame de ruts_con_problema).
    """
    if "Rut" not in df_empleados.columns:
        return set(), pd.DataFrame()

    conteo = df_empleados.groupby("Rut").size().reset_index(name="n_contratos")
    problemas = conteo[conteo["n_contratos"] > 1].copy()

    if not problemas.empty:
        # Agregar datos adicionales del empleado para el log
        detalle = df_empleados[df_empleados["Rut"].isin(problemas["Rut"])][
            [c for c in ["Rut", "Nombre", "Empresa", "Contrato"] if c in df_empleados.columns]
        ].copy()
        detalle = detalle.merge(problemas, on="Rut", how="left")
        return set(problemas["Rut"].tolist()), detalle
    else:
        return set(), pd.DataFrame()


# ─────────────────────────────────────────────
# LOOKUP HELPERS
# ─────────────────────────────────────────────
def lookup(df, col_busca, valor, col_trae, default=""):
    """Busca valor en col_busca y retorna col_trae."""
    if df is None or df.empty:
        return default
    try:
        valor_num = pd.to_numeric(valor, errors="coerce")
        mask = df[col_busca] == (valor_num if not pd.isna(valor_num) else valor)
        if mask.any():
            return df.loc[mask.idxmax(), col_trae]
    except Exception:
        pass
    return default


def safe_num(val, default=0):
    """Convierte a número de forma segura."""
    try:
        v = float(val)
        return v if not np.isnan(v) else default
    except Exception:
        return default


# ─────────────────────────────────────────────
# GENERACIÓN DE FILAS DE SALIDA
# ─────────────────────────────────────────────
def generar_filas_dt(df, fecha_proceso, refs, df_empleados):
    """
    Genera las filas del archivo de salida desde el CSV de la DT.
    """
    filas = []

    equiv        = refs.get("equiv_conceptos", pd.DataFrame())
    params_df    = refs.get("parametros", pd.DataFrame())
    inst_afp     = refs.get("inst_afp", pd.DataFrame())
    inst_mutuales= refs.get("inst_mutuales", pd.DataFrame())
    inst_salud   = refs.get("inst_salud", pd.DataFrame())
    inst_cajas   = refs.get("inst_cajas", pd.DataFrame())
    empresas_raw = refs.get("listado_empresas", pd.DataFrame())

    # Cargar empresas con encabezado correcto
    df_empresas = cargar_empresas(empresas_raw)

    # ── Parámetros del mes ──
    tope_salud = 0
    tope_afp   = 0
    tope_ces   = 0
    if not params_df.empty and "mes_Proc" in params_df.columns:
        params_df["mes_Proc"] = params_df["mes_Proc"].astype(str).str.strip()
        row_params = params_df[params_df["mes_Proc"] == fecha_proceso]
        if not row_params.empty:
            tope_salud = safe_num(row_params.iloc[0].get("topeSalud_pesos", 0))
            tope_afp   = safe_num(row_params.iloc[0].get("topeImp_pesos_afp", 0))
            tope_ces   = safe_num(row_params.iloc[0].get("topeCes_pesos", 0))

    # ── Mapa de equivalencias: cod_lre → concepto_detalle + Tipo ──
    equiv_map  = {}  # col_csv → concepto_detalle
    tipo_map   = {}  # concepto_detalle → Tipo
    if not equiv.empty and "cod_lre" in equiv.columns and "concepto_detalle" in equiv.columns:
        for _, er in equiv.iterrows():
            col_csv   = str(er["cod_lre"]).strip()
            concepto  = str(er["concepto_detalle"]).strip()
            tipo      = str(er.get("Tipo", "")).strip()
            # Solo mapear la primera aparición para evitar duplicados de isapre
            if col_csv not in equiv_map:
                equiv_map[col_csv] = concepto
            tipo_map[concepto] = tipo

    # Columnas del CSV que tienen equivalencia (excluir COL_SALUD_VOL para isapre, se suma manualmente)
    cols_conceptos = [c for c in df.columns if c in equiv_map and c != COL_SALUD_VOL]

    # ── Detectar múltiples contratos ──
    ruts_multiples, _ = detectar_multiples_contratos(df_empleados)

    for _, row in df.iterrows():
        rut = str(row.get(COL_RUT, "")).strip()

        # Excluir trabajadores con múltiples contratos
        if rut in ruts_multiples:
            continue

        # ── Lookup empleado → empresa ──
        empresa_codigo = ""
        empresa_salida = ""
        numero_contrato = ""
        if not df_empleados.empty and "Rut" in df_empleados.columns:
            emp_rows = df_empleados[df_empleados["Rut"] == rut]
            if not emp_rows.empty:
                empresa_codigo = str(emp_rows.iloc[0].get("Empresa", "")).strip()
                if "Contrato" in df_empleados.columns:
                    numero_contrato = emp_rows.iloc[0].get("Contrato", "")

        # ── Lookup empresa → código empresa ──
        if empresa_codigo and not df_empresas.empty and "Empresa" in df_empresas.columns:
            emp2 = df_empresas[df_empresas["Empresa"] == empresa_codigo]
            if not emp2.empty:
                empresa_salida = emp2.iloc[0]["Empresa"]
            else:
                empresa_salida = empresa_codigo  # fallback al código directo

        # ── Valores base del trabajador ──
        dias_trab   = safe_num(row.get(COL_DIAS_TRAB, 0))
        dias_lic    = safe_num(row.get(COL_DIAS_LIC, 0))
        dias_vac    = safe_num(row.get(COL_DIAS_VAC, 0))
        sueldo      = safe_num(row.get(COL_SUELDO, 0))
        rebaja_zona = safe_num(row.get(COL_REBAJA_ZONA, 0))

        monto_init = round((sueldo / dias_trab) * 30, 0) if dias_trab > 0 else 0

        # ── Suma haberes afectos (para Afecto) ──
        conceptos_haber_afecto = [c for c in cols_conceptos
                                   if tipo_map.get(equiv_map.get(c, ""), "") == "Haber afecto"]
        suma_haber_afecto = sum(safe_num(row.get(c, 0)) for c in conceptos_haber_afecto)

        # ── Suma haberes exentos (para Rentas no gravadas) ──
        conceptos_haber_exento = [c for c in cols_conceptos
                                   if tipo_map.get(equiv_map.get(c, ""), "") == "Haber exento"]
        suma_haber_exento = sum(safe_num(row.get(c, 0)) for c in conceptos_haber_exento)

        # ── Monto isapre (caso especial: suma dos columnas) ──
        monto_isapre = safe_num(row.get(COL_SALUD7, 0)) + safe_num(row.get(COL_SALUD_VOL, 0))

        # ── Total rebajas por LLSS (solo para concepto impuesto) ──
        salud_trab = safe_num(row.get(COL_SALUD7, 0)) + safe_num(row.get(COL_SALUD_VOL, 0))
        total_rebajas_llss = (
            safe_num(row.get(COL_AFP, 0))
            + safe_num(row.get(COL_CES_TRAB, 0))
            + safe_num(row.get(COL_APVI_MOD_B, 0))
            + safe_num(row.get(COL_TRAB_PESADO, 0))
            + min(salud_trab, tope_salud) if tope_salud > 0 else salud_trab
        )

        # ── Código de institución del trabajador ──
        cod_afp_inst    = safe_num(row.get(COL_AFP_INST, 0))
        cod_isapre_inst = safe_num(row.get(COL_ISAPRE_INST, 0))
        cod_mutual_inst = safe_num(row.get(COL_MUTUAL_INST, 0))
        cod_ccaf_inst   = safe_num(row.get(COL_CCAF_INST, 0))

        # ── Lookup instituciones ──
        id_afp_trab    = lookup(inst_afp,      "cod_lre", cod_afp_inst,    "id_afp",    "")
        id_isapre_trab = lookup(inst_salud,    "cod_lre", cod_isapre_inst, "id_inst",   "")
        id_mutual_trab = lookup(inst_mutuales, "cod_lre", cod_mutual_inst, "id_mutual", "")
        id_ccaf_trab   = lookup(inst_cajas,    "cod_lre", cod_ccaf_inst,   "id_ccaf",   "")
        cot_afp_trab   = lookup(inst_afp,      "cod_lre", cod_afp_inst,    "cot_afp",   0)

        # ── Cotización mutual (triangulación empleado → empresa) ──
        cot_mutual = 0.93
        if empresa_codigo and not df_empresas.empty and "Empresa" in df_empresas.columns:
            emp2 = df_empresas[df_empresas["Empresa"] == empresa_codigo]
            if not emp2.empty and "Cotización Mutual" in df_empresas.columns:
                cot_mutual = safe_num(emp2.iloc[0].get("Cotización Mutual", 0.93), 0.93)

        # ── Generar fila por cada concepto ──
        for col_csv in cols_conceptos:
            id_concepto = equiv_map.get(col_csv, "")
            if not id_concepto:
                continue

            # Monto especial para isapre
            if id_concepto == "isapre":
                monto = monto_isapre
            else:
                monto = safe_num(row.get(col_csv, 0))

            # Saltar si monto es 0 (no hay movimiento)
            if monto == 0:
                continue

            # ── Id de institución ──
            id_institucion = ""
            if id_concepto in CONCEPTOS_ID_AFP:
                id_institucion = id_afp_trab
            elif id_concepto == "isapre":
                id_institucion = id_isapre_trab
            elif id_concepto == "mutual":
                id_institucion = id_mutual_trab
            elif id_concepto == "cajaCred":
                id_institucion = id_ccaf_trab

            # ── Afecto ──
            if id_concepto in CONCEPTOS_AFECTO_AFP:
                afecto = min(suma_haber_afecto, tope_afp) if tope_afp > 0 else suma_haber_afecto
            elif id_concepto in CONCEPTOS_AFECTO_CES:
                afecto = min(suma_haber_afecto, tope_ces) if tope_ces > 0 else suma_haber_afecto
            else:
                afecto = 0

            # ── Cotización de jubilación ──
            cot_jubilacion = ""
            if id_concepto == "cesEmpleado":
                cot_jubilacion = 0.6
            elif id_concepto == "afp":
                cot_jubilacion = cot_afp_trab
            elif id_concepto == "isapre" and id_institucion == "fonasa":
                cot_jubilacion = monto
            elif id_concepto == "mutual":
                cot_jubilacion = cot_mutual

            # ── Rentas no gravadas (solo si concepto = impuesto) ──
            rentas_no_grav = suma_haber_exento if id_concepto == "impuesto" else 0

            # ── Total rebajas LLSS (solo si concepto = impuesto) ──
            rebajas_llss = total_rebajas_llss if id_concepto == "impuesto" else 0

            filas.append({
                "Fecha de proceso":        fecha_proceso,
                "Id empleado":             rut,
                "Número de contrato":      numero_contrato,
                "Id del concepto":         id_concepto,
                "Monto del concepto":      monto,
                "Afecto":                  afecto,
                "Id de institución":       id_institucion,
                "Cotización de jubilación": cot_jubilacion,
                "Días de licencias":       dias_lic,
                "Días trabajados":         dias_trab,
                "Fecha de aplicación":     fecha_proceso,
                "Empresa":                 empresa_salida,
                "Total de rebajas por LLSS": rebajas_llss,
                "Rentas no gravadas":      rentas_no_grav,
                "Rebaja por zona extrema": rebaja_zona,
                "Jornada":                 "C",
                "Días de vacaciones":      dias_vac,
                "Monto Init":              monto_init,
                "Fase":                    1,
            })

    return pd.DataFrame(filas)


# ─────────────────────────────────────────────
# GENERACIÓN EXCEL DE SALIDA
# ─────────────────────────────────────────────
def generar_excel_dt(df_salida):
    """Genera el Excel de salida con formato Rex+."""
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Liquidaciones"

    header_fill = PatternFill("solid", fgColor="1A2744")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(
        bottom=Side(style="thin", color="E8EDF5"),
        right=Side(style="thin", color="E8EDF5")
    )

    cols = list(df_salida.columns)
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = max(len(col) + 4, 14)

    for ri, row in enumerate(df_salida.itertuples(index=False), 2):
        fill = PatternFill("solid", fgColor="EAF0F8") if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    wb.save(output)
    return output.getvalue()


# ─────────────────────────────────────────────
# GENERACIÓN EXCEL LOG MÚLTIPLES CONTRATOS
# ─────────────────────────────────────────────
def generar_excel_log(df_log):
    """Genera el Excel de log de trabajadores con múltiples contratos."""
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Múltiples contratos"

    header_fill = PatternFill("solid", fgColor="C53030")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    cols = list(df_log.columns)
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = max(len(col) + 4, 18)

    for ri, row in enumerate(df_log.itertuples(index=False), 2):
        fill = PatternFill("solid", fgColor="FFF5F5") if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    wb.save(output)
    return output.getvalue()


# ─────────────────────────────────────────────
# INTERFAZ STREAMLIT DEL MÓDULO DT
# ─────────────────────────────────────────────
def render_modulo_dt(refs_compartidas):
    """
    Renderiza la interfaz del módulo DT dentro del tab de Streamlit.
    refs_compartidas: dict con archivos de referencia ya cargados por app_migracion.
    """
    st.markdown('<div class="section-title">🏛️ Migración Declaración Jurada DT</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-sub">Procesa el archivo descargado directamente desde la Dirección del Trabajo y genera el archivo de liquidaciones en detalle Rex+.</div>', unsafe_allow_html=True)

    # Pasos
    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown("""
        <div class="step-card">
            <div class="step-label">PASO 1</div>
            <div class="step-title">Subir archivo DT</div>
            <div class="step-desc">Archivo CSV descargado desde el portal de la Dirección del Trabajo.</div>
        </div>""", unsafe_allow_html=True)
    with col2:
        st.markdown("""
        <div class="step-card">
            <div class="step-label">PASO 2</div>
            <div class="step-title">Subir listado de empleados</div>
            <div class="step-desc">Listado de empleados del período a procesar exportado desde Rex+.</div>
        </div>""", unsafe_allow_html=True)
    with col3:
        st.markdown("""
        <div class="step-card">
            <div class="step-label">PASO 3</div>
            <div class="step-title">Descargar resultado</div>
            <div class="step-desc">Se genera el archivo Excel de liquidaciones listo para importar.</div>
        </div>""", unsafe_allow_html=True)

    st.markdown('<hr class="rex-divider">', unsafe_allow_html=True)

    # ── Upload archivo DT ──
    st.markdown("### 📤 Archivo CSV de la Dirección del Trabajo")
    archivo_dt = st.file_uploader(
        "Selecciona el archivo CSV de la DT",
        type=["csv"],
        accept_multiple_files=False,
        key="dt_csv_upload",
        help="Archivo descargado desde el portal de la Dirección del Trabajo. El encabezado puede estar al inicio o al final."
    )

    # ── Upload listado empleados ──
    st.markdown("### 👥 Listado de empleados del período")
    archivo_empleados = st.file_uploader(
        "Sube el archivo listado_empleados.xlsx del período",
        type=["xlsx"],
        accept_multiple_files=False,
        key="dt_empleados_upload",
        help="Exportado desde Rex+. Debe contener columnas: Rut, Nombre, Empresa, Contrato."
    )

    if archivo_empleados:
        st.markdown(f'<div class="alert-success">✅ Listado de empleados cargado: <b>{archivo_empleados.name}</b></div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="alert-warning">⚠️ Debes subir el listado de empleados para ejecutar el proceso.</div>', unsafe_allow_html=True)

    if not archivo_dt or not archivo_empleados:
        return

    st.markdown(f'<div class="alert-success">✅ Archivo DT cargado: <b>{archivo_dt.name}</b></div>', unsafe_allow_html=True)

    # ── Determinar fecha de proceso ──
    fecha_proceso, fecha_ok = extraer_fecha_dt(archivo_dt.name)

    if not fecha_ok:
        st.markdown("""
        <div class="alert-warning">
            ⚠️ <b>No se pudo determinar la fecha de proceso</b> desde el nombre del archivo.<br>
            Por favor selecciona el mes y año correspondiente.
        </div>""", unsafe_allow_html=True)

        col_m, col_a, _ = st.columns([1, 1, 3])
        with col_m:
            mes_sel = st.selectbox("Mes", options=list(range(1, 13)),
                                   format_func=lambda x: f"{x:02d} - {['Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre'][x-1]}",
                                   key="dt_mes_sel")
        with col_a:
            anio_actual = datetime.now().year
            anio_sel = st.selectbox("Año", options=list(range(anio_actual - 5, anio_actual + 2)),
                                    index=5, key="dt_anio_sel")
        fecha_proceso = f"{anio_sel}-{mes_sel:02d}"

    st.markdown(f'<div class="alert-success">📅 Fecha de proceso: <b>{fecha_proceso}</b></div>', unsafe_allow_html=True)

    # ── Verificar parámetros del mes ──
    params_df = refs_compartidas.get("parametros", pd.DataFrame())
    if not params_df.empty and "mes_Proc" in params_df.columns:
        params_df["mes_Proc"] = params_df["mes_Proc"].astype(str).str.strip()
        if fecha_proceso not in params_df["mes_Proc"].values:
            st.markdown(f"""
            <div class="alert-error">
                ❌ <b>No se encontraron parámetros mensuales para {fecha_proceso}.</b><br>
                Ve a la pestaña <b>⚙️ Parámetros Mensuales</b> y agrega el mes antes de continuar.
            </div>""", unsafe_allow_html=True)
            return

    # ── Botón ejecutar ──
    if st.button("▶ Ejecutar proceso DT", key="dt_btn_ejecutar"):
        with st.spinner("Procesando archivo..."):
            try:
                # Leer CSV DT
                df_dt = leer_csv_dt(archivo_dt)

                # Leer empleados
                df_empleados = cargar_empleados(archivo_empleados)

                # Detectar múltiples contratos
                ruts_multiples, df_log_mult = detectar_multiples_contratos(df_empleados)

            except Exception as e:
                st.markdown(f'<div class="alert-error">❌ Error al leer los archivos: <b>{e}</b></div>', unsafe_allow_html=True)
                return

        st.markdown('<hr class="rex-divider">', unsafe_allow_html=True)
        st.markdown("### 🔍 Resultado del proceso")

        # ── Mostrar log de múltiples contratos si hay ──
        if ruts_multiples:
            st.markdown(f"""
            <div class="alert-warning">
                ⚠️ <b>{len(ruts_multiples)} trabajador(es) con múltiples contratos</b> fueron excluidos del archivo de salida.<br>
                Descarga el log para revisarlos.
            </div>""", unsafe_allow_html=True)

            with st.expander(f"👁️ Ver trabajadores excluidos ({len(ruts_multiples)})"):
                st.dataframe(df_log_mult, use_container_width=True, hide_index=True)

            log_bytes = generar_excel_log(df_log_mult)
            st.download_button(
                label="⬇️ Descargar log_multiples_contratos.xlsx",
                data=log_bytes,
                file_name=f"log_multiples_contratos_{fecha_proceso}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dt_btn_log"
            )

        # ── Generar archivo de salida ──
        with st.spinner("Generando archivo de salida..."):
            try:
                df_salida = generar_filas_dt(df_dt, fecha_proceso, refs_compartidas, df_empleados)
            except Exception as e:
                st.markdown(f'<div class="alert-error">❌ Error al generar el archivo de salida: <b>{e}</b></div>', unsafe_allow_html=True)
                return

        if df_salida.empty:
            st.markdown('<div class="alert-warning">⚠️ No se generaron registros. Verifica que el archivo y los parámetros sean correctos.</div>', unsafe_allow_html=True)
            return

        st.markdown(f"""
        <div class="alert-success">
            ✅ <b>Archivo generado exitosamente.</b><br>
            Se procesaron <b>{df_salida["Id empleado"].nunique()} trabajadores</b> y se generaron <b>{len(df_salida)} registros</b>.
        </div>""", unsafe_allow_html=True)

        with st.expander("👁️ Vista previa del archivo de salida"):
            st.dataframe(df_salida.head(50), use_container_width=True, hide_index=True)

        excel_bytes = generar_excel_dt(df_salida)
        st.download_button(
            label="⬇️ Descargar archivo de salida (.xlsx)",
            data=excel_bytes,
            file_name=f"liquidaciones_dt_{fecha_proceso}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dt_btn_descarga"
        )
