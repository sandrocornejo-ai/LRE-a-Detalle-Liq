import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import io
from datetime import datetime

# ─────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────
DATA_DIR = "data"  # Carpeta con archivos de referencia

# ─────────────────────────────────────────────
# ESTILOS REX+
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Rex+ | Migración DDJJ Previred",
    page_icon="💼",
    layout="wide"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

html, body, [class*="css"] {
    font-family: 'Inter', sans-serif;
}

/* Header */
.rex-header {
    background-color: #1a2744;
    padding: 14px 28px;
    border-radius: 10px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    margin-bottom: 28px;
}
.rex-logo {
    background: white;
    color: #1a2744;
    font-weight: 800;
    font-size: 15px;
    padding: 5px 10px;
    border-radius: 6px;
    letter-spacing: 0.5px;
}
.rex-logo span { color: #00b4d8; }
.rex-title { color: white; font-size: 18px; font-weight: 600; margin-left: 16px; }
.rex-badge {
    background: #00b4d8;
    color: white;
    font-size: 11px;
    font-weight: 700;
    padding: 4px 12px;
    border-radius: 20px;
    letter-spacing: 1px;
}

/* Cards */
.step-card {
    background: white;
    border: 1px solid #e8edf5;
    border-radius: 10px;
    padding: 18px 20px;
    margin-bottom: 12px;
}
.step-label {
    color: #00b4d8;
    font-size: 11px;
    font-weight: 700;
    letter-spacing: 1px;
    text-transform: uppercase;
    margin-bottom: 4px;
}
.step-title { font-size: 15px; font-weight: 600; color: #1a2744; margin-bottom: 4px; }
.step-desc { font-size: 13px; color: #6b7a9a; }

/* Section title */
.section-title {
    font-size: 20px;
    font-weight: 700;
    color: #1a2744;
    margin-bottom: 4px;
}
.section-sub { font-size: 13px; color: #6b7a9a; margin-bottom: 20px; }

/* Alerts */
.alert-error {
    background: #fff0f0;
    border-left: 4px solid #e53e3e;
    border-radius: 6px;
    padding: 12px 16px;
    margin: 8px 0;
    font-size: 13px;
    color: #c53030;
}
.alert-success {
    background: #f0fff4;
    border-left: 4px solid #38a169;
    border-radius: 6px;
    padding: 12px 16px;
    margin: 8px 0;
    font-size: 13px;
    color: #276749;
}
.alert-warning {
    background: #fffbf0;
    border-left: 4px solid #d69e2e;
    border-radius: 6px;
    padding: 12px 16px;
    margin: 8px 0;
    font-size: 13px;
    color: #744210;
}

/* Divider */
.rex-divider {
    border: none;
    border-top: 1px solid #e8edf5;
    margin: 24px 0;
}

/* Log table */
.log-header {
    background: #1a2744;
    color: white;
    font-size: 13px;
    font-weight: 600;
    padding: 10px 14px;
    border-radius: 8px 8px 0 0;
}

/* Buttons override */
.stButton > button {
    background-color: #1a2744 !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    padding: 8px 20px !important;
    font-size: 14px !important;
}
.stButton > button:hover {
    background-color: #00b4d8 !important;
    color: white !important;
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# CONSTANTES DE COLUMNAS
# ─────────────────────────────────────────────
COLS_HABERES_AFECTOS = [
    "Sueldo(2101)", "Sobresueldo(2102)", "Comisiones(2103)", "Semana corrida(2104)",
    "Participación(2105)", "Gratificación(2106)", "Recargo 30% día domingo(2107)",
    "Remun. variable pagada en vacaciones(2108)", "Remun. variable pagada en clausura(2109)",
    "Aguinaldo(2110)", "Bonos u otras remun. fijas mensuales(2111)", "Tratos(2112)",
    "Bonos u otras remun. variables mensuales o superiores a un mes(2113)",
    "Ejercicio opción no pactada en contrato(2114)",
    "Beneficios en especie constitutivos de remun(2115)",
    "Remuneraciones bimestrales(2116)", "Remuneraciones trimestrales(2117)",
    "Remuneraciones cuatrimestral(2118)", "Remuneraciones semestrales(2119)",
    "Remuneraciones anuales(2120)", "Participación anual(2121)",
    "Gratificación anual(2122)", "Otras remuneraciones superiores a un mes(2123)",
    "Pago por horas de trabajo sindical(2124)", "Sueldo empresarial (2161)",
    "Subsidio por incapacidad laboral por licencia médica(2201)",
    "Beca de estudio(2202)", "Gratificaciones de zona(2203)"
]

COLS_HABERES_EXENTOS = [
    "Otros ingresos no constitutivos de renta(2204)", "Colación(2301)",
    "Movilización(2302)", "Viáticos(2303)", "Asignación de pérdida de caja(2304)",
    "Asignación de desgaste herramienta(2305)", "Asignación familiar legal(2311)",
    "Gastos por causa del trabajo(2306)", "Gastos por cambio de residencia(2307)",
    "Sala cuna(2308)", "Asignación trabajo a distancia o teletrabajo(2309)",
    "Depósito convenido hasta UF 900(2347)", "Alojamiento por razones de trabajo(2310)",
    "Asignación de traslación(2312)", "Indemnización por feriado legal(2313)",
    "Indemnización años de servicio(2314)", "Indemnización sustitutiva del aviso previo(2315)",
    "Indemnización fuero maternal(2316)", "Pago indemnización a todo evento(2331)",
    "Indemnizaciones voluntarias tributables(2417)",
    "Indemnizaciones contractuales tributables(2418)"
]

COLS_DESCUENTOS_LEGALES = [
    "Cotización obligatoria previsional (AFP o IPS)(3141)",
    "Cotización obligatoria salud 7%(3143)", "Cotización voluntaria para salud(3144)",
    "Cotización AFC - trabajador(3151)",
    "Cotizaciones técnico extranjero para seguridad social fuera de Chile(3146)",
    "Descuento depósito convenido hasta UF 900 anual(3147)",
    "Cotización APVi Mod A(3155)", "Cotización APVi Mod B hasta UF50(3156)",
    "Cotización APVc Mod A(3157)", "Cotización APVc Mod B hasta UF50(3158)",
    "Impuesto retenido por remuneraciones(3161)",
    "Impuesto retenido por indemnizaciones(3162)",
    "Mayor retención de impuestos solicitada por el trabajador(3163)",
    "Impuesto retenido por reliquidación remun. devengadas otros períodos(3164)",
    "Diferencia impuesto reliquidación remun. devengadas en este período(3165)",
    "Retención préstamo clase media 2020 (Ley 21.252) (3166)",
    "Rebaja zona extrema DL 889 (3167)"
]

COLS_OTROS_DESCUENTOS = [
    "Cuota sindical 1(3171)", "Cuota sindical 2(3172)", "Cuota sindical 3(3173)",
    "Cuota sindical 4(3174)", "Cuota sindical 5(3175)", "Cuota sindical 6(3176)",
    "Cuota sindical 7(3177)", "Cuota sindical 8(3178)", "Cuota sindical 9(3179)",
    "Cuota sindical 10(3180)", "Crédito social CCAF(3110)",
    "Cuota vivienda o educación(3181)", "Crédito cooperativas de ahorro(3182)",
    "Otros descuentos autorizados y solicitados por el trabajador(3183)",
    "Cotización adicional trabajo pesado - trabajador(3154)",
    "Donaciones culturales y de reconstrucción(3184)", "Otros descuentos(3185)",
    "Pensiones de alimentos(3186)", "Descuento mujer casada(3187)",
    "Descuentos por anticipos y préstamos(3188)"
]

COLS_APORTES_EMPLEADOR = [
    "AFC - Aporte empleador(4151)",
    "Aporte empleador seguro accidentes del trabajo y Ley SANNA(4152)",
    "Aporte empleador indemnización a todo evento(4131)",
    "Aporte adicional trabajo pesado - empleador(4154)",
    "Aporte empleador seguro invalidez y sobrevivencia(4155)",
    "APVC - Aporte Empleador(4157)"
]

GRUPOS_AFP = {
    "afp", "reliquidaAfp", "afpAhor", "cesEmpleado", "reliquidaCesEmpl",
    "trabajoPesaEmpl", "voluntarioCoti", "voluntarioAhor", "reliquidaTrabEmpl",
    "trabajoPesa", "reliquidaTrabPesa", "sis", "reliquidaSis", "cesAporteCi",
    "reliquidaCesCi", "cesAporteSol", "reliquidaCesSol", "aporteAFPemp"
}
GRUPOS_ISAPRE = {"isapre", "reliquidaIsapre"}
GRUPOS_MUTUAL = {"mutual", "reliquidaMutual"}
GRUPOS_CCAF = {
    "ccafReliquida", "cajaCred", "cajaDent", "cajaLeas", "cajaVida",
    "cajaOtro", "cajaAhor", "cajaSegu", "cajaComp", "reliquidaCcaf"
}
GRUPOS_AFP_MUTUAL_AFECTO = {
    "afp", "isapre", "reliquidaIsapre", "reliquidaAfp", "afpAhor",
    "trabajoPesaEmpl", "voluntarioCoti", "voluntarioAhor", "reliquidaTrabEmpl",
    "trabajoPesa", "reliquidaTrabPesa", "sis", "reliquidaSis", "aporteAFPemp",
    "mutual", "reliquidaMutual"
}
GRUPOS_CES_AFECTO = {
    "cesEmpleado", "reliquidaCesEmpl", "cesAporteCi",
    "reliquidaCesCi", "cesAporteSol", "reliquidaCesSol"
}

# ─────────────────────────────────────────────
# FUNCIONES DE CARGA DE REFERENCIAS
# ─────────────────────────────────────────────
@st.cache_data
def cargar_referencias():
    refs = {}
    archivos = {
        "equiv_conceptos": "equiv_conceptos.xlsx",
        "listado_empleados": "listado_empleados.xlsx",
        "listado_empresas": "listado_empresas.xlsx",
        "inst_mutuales": "inst_mutuales.xlsx",
        "inst_cajas": "inst_cajas.xlsx",
        "cot_afp_hist": "cot_afp_hist.xlsx",
        "parametros": "parametrosMensuales.xlsx",
    }
    errores = []
    for key, fname in archivos.items():
        path = os.path.join(DATA_DIR, fname)
        if os.path.exists(path):
            refs[key] = pd.read_excel(path)
        else:
            errores.append(fname)
    return refs, errores

# ─────────────────────────────────────────────
# FUNCIONES DE PROCESAMIENTO
# ─────────────────────────────────────────────
def extraer_fecha_proceso(nombre_archivo):
    """Extrae los últimos 6 chars antes de .csv → formato aaaa-mm"""
    base = os.path.splitext(nombre_archivo)[0]
    sufijo = base[-6:]
    return f"{sufijo[:4]}-{sufijo[4:]}"

def safe_sum(df, cols):
    """Suma columnas que existen en el df, ignorando las que no existen."""
    cols_presentes = [c for c in cols if c in df.columns]
    if not cols_presentes:
        return pd.Series(0, index=df.index)
    return df[cols_presentes].fillna(0).sum(axis=1)

def get_col(df, col, default=0):
    """Obtiene una columna del df o retorna default si no existe."""
    if col in df.columns:
        return df[col].fillna(default)
    return pd.Series(default, index=df.index)

def validar_archivos(archivos_subidos):
    """Valida que todos los archivos sean de la misma empresa."""
    nombres = [f.name for f in archivos_subidos]
    prefijos = [n[:10] for n in nombres]
    if len(set(prefijos)) > 1:
        return False, prefijos
    return True, prefijos

def calcular_totales(df):
    """Calcula los 5 totales pre-validación."""
    df = df.copy()
    df["_total_haberes_afectos"] = safe_sum(df, COLS_HABERES_AFECTOS)
    df["_total_haberes_exentos"] = safe_sum(df, COLS_HABERES_EXENTOS)
    df["_total_descuentos_legales"] = safe_sum(df, COLS_DESCUENTOS_LEGALES)
    df["_total_otros_descuentos"] = safe_sum(df, COLS_OTROS_DESCUENTOS)
    df["_total_aportes_empleador"] = safe_sum(df, COLS_APORTES_EMPLEADOR)
    return df

def validar_cuadraturas(df, nombre_archivo):
    """Ejecuta las 6 validaciones y retorna lista de errores."""
    errores = []
    tol = 1  # tolerancia de 1 peso por redondeo

    validaciones = [
        ("V1", "_total_haberes_afectos", "Total haberes imponibles y tributables(5210)",
         "total_haberes_afectos ≠ Total haberes imponibles y tributables(5210)"),
        ("V2", "_total_haberes_exentos", "Total haberes no imponibles y no tributables(5230)",
         "total_haberes_exentos ≠ Total haberes no imponibles y no tributables(5230)"),
        ("V3", "_total_descuentos_legales", "Total descuentos por cotizaciones del trabajador(5341)",
         "total_descuentos_legales ≠ Total descuentos por cotizaciones del trabajador(5341)"),
        ("V4", "_total_otros_descuentos", "Total otros descuentos(5302)",
         "total_otros_descuentos ≠ Total otros descuentos(5302)"),
        ("V5", "_total_aportes_empleador", "Total aportes empleador(5410)",
         "total_aportes_empleador ≠ Total aportes empleador(5410)"),
    ]

    for codigo, col_calc, col_ctrl, mensaje in validaciones:
        if col_ctrl not in df.columns:
            continue
        ctrl = df[col_ctrl].fillna(0)
        calc = df[col_calc].fillna(0)
        mask = (calc - ctrl).abs() > tol
        filas_error = df[mask]
        for _, row in filas_error.iterrows():
            errores.append({
                "Archivo": nombre_archivo,
                "RUT": row.get("Rut trabajador (1101)", "N/D"),
                "Validación": codigo,
                "Descripción": mensaje,
                "Valor calculado": round(calc[row.name], 2),
                "Valor control": round(ctrl[row.name], 2),
                "Diferencia": round(calc[row.name] - ctrl[row.name], 2)
            })

    # V6: liquidez
    if all(c in df.columns for c in ["Total líquido(5501)"]):
        liq_calc = (df["_total_haberes_afectos"] + df["_total_haberes_exentos"]) - \
                   (df["_total_descuentos_legales"] + df["_total_otros_descuentos"])
        liq_ctrl = df["Total líquido(5501)"].fillna(0)
        mask = (liq_calc - liq_ctrl).abs() > tol
        filas_error = df[mask]
        for _, row in filas_error.iterrows():
            errores.append({
                "Archivo": nombre_archivo,
                "RUT": row.get("Rut trabajador (1101)", "N/D"),
                "Validación": "V6",
                "Descripción": "(haberes_afectos + haberes_exentos) - (desc_legales + otros_desc) ≠ Total líquido(5501)",
                "Valor calculado": round(liq_calc[row.name], 2),
                "Valor control": round(liq_ctrl[row.name], 2),
                "Diferencia": round(liq_calc[row.name] - liq_ctrl[row.name], 2)
            })

    return errores

def generar_filas_salida(df, fecha_proceso, refs):
    """Genera las filas del archivo de salida via pivot de conceptos."""
    filas = []

    # Cargar lookups
    equiv = refs.get("equiv_conceptos", pd.DataFrame())
    empleados = refs.get("listado_empleados", pd.DataFrame())
    empresas = refs.get("listado_empresas", pd.DataFrame())
    mutuales = refs.get("inst_mutuales", pd.DataFrame())
    cajas = refs.get("inst_cajas", pd.DataFrame())
    cot_afp = refs.get("cot_afp_hist", pd.DataFrame())
    params = refs.get("parametros", pd.DataFrame())

    # Diccionario de equivalencias de conceptos
    equiv_dict = {}
    if not equiv.empty and "cod_lre" in equiv.columns and "concepto_detalle" in equiv.columns:
        equiv_dict = dict(zip(equiv["cod_lre"], equiv["concepto_detalle"]))

    # Parámetros mensuales (primera fila)
    tope_afp = 0
    tope_ces = 0
    if not params.empty:
        if "topeImp_pesos_afp" in params.columns:
            tope_afp = params["topeImp_pesos_afp"].iloc[0]
        if "topeCes_pesos" in params.columns:
            tope_ces = params["topeCes_pesos"].iloc[0]

    # Columnas de conceptos (las que están en equiv_dict)
    cols_concepto = [c for c in df.columns if c in equiv_dict]

    for _, row in df.iterrows():
        rut = row.get("Rut trabajador (1101)", "")

        # Lookup empresa
        empresa_salida = ""
        if not empleados.empty and "Rut" in empleados.columns:
            emp_row = empleados[empleados["Rut"] == rut]
            if not emp_row.empty:
                nombre_empresa = emp_row.iloc[0].get("Empresa", "")
                if not empresas.empty and "Nombre" in empresas.columns:
                    emp2 = empresas[empresas["Nombre"] == nombre_empresa]
                    if not emp2.empty:
                        empresa_salida = emp2.iloc[0].iloc[0]

        # AFP e Isapre del empleado
        afp_empleado = ""
        isapre_empleado = ""
        if not empleados.empty and "Rut" in empleados.columns:
            emp_row = empleados[empleados["Rut"] == rut]
            if not emp_row.empty:
                afp_empleado = emp_row.iloc[0].get("AFP", "")
                isapre_empleado = emp_row.iloc[0].get("Isapre", "")

        dias_trabajados = row.get("Nro días trabajados en el mes(1115)", 0) or 0
        dias_licencia = row.get("Nro días de licencia médica en el mes(1116)", 0) or 0
        dias_vacaciones = row.get("Nro días de vacaciones en el mes(1117)", 0) or 0
        sueldo = row.get("Sueldo(2101)", 0) or 0
        total_imponible = row.get("Total haberes imponibles y tributables(5210)", 0) or 0
        col_1152 = row.get("Org. administrador ley 16.744(1152)", "")
        col_3110 = row.get("Crédito social CCAF(3110)", 0) or 0
        rebaja_zona = row.get("Rebaja zona extrema DL 889 (3167)", 0) or 0

        monto_init = (sueldo / dias_trabajados * 30) if dias_trabajados > 0 else 0

        # Fila por cada concepto
        for col_csv in cols_concepto:
            monto = row.get(col_csv, 0) or 0
            id_concepto = equiv_dict.get(col_csv, "")

            # Id de institución
            id_institucion = ""
            if id_concepto in GRUPOS_AFP:
                id_institucion = afp_empleado
            elif id_concepto in GRUPOS_ISAPRE:
                id_institucion = isapre_empleado
            elif id_concepto in GRUPOS_MUTUAL:
                if not mutuales.empty and "cod_lre" in mutuales.columns and "id_mutual" in mutuales.columns:
                    m = mutuales[mutuales["cod_lre"] == col_1152]
                    if not m.empty:
                        id_institucion = m.iloc[0]["id_mutual"]
            elif id_concepto in GRUPOS_CCAF and col_3110 != 0:
                if not cajas.empty:
                    c = cajas[cajas.iloc[:, 0] == col_3110]
                    if not c.empty and "id_ccaf" in cajas.columns:
                        id_institucion = c.iloc[0]["id_ccaf"]

            # Afecto
            afecto = ""
            if id_concepto in GRUPOS_AFP_MUTUAL_AFECTO:
                afecto = min(total_imponible, tope_afp) if tope_afp > 0 else total_imponible
            elif id_concepto in GRUPOS_CES_AFECTO:
                afecto = min(total_imponible, tope_ces) if tope_ces > 0 else total_imponible

            # Cotización de jubilación
            cot_jubilacion = 0
            if id_concepto == "afp":
                key_afp = f"{fecha_proceso}{id_institucion}"
                if not cot_afp.empty and "id_afp_hist" in cot_afp.columns:
                    r = cot_afp[cot_afp["id_afp_hist"] == key_afp]
                    if not r.empty:
                        cot_jubilacion = r.iloc[0].get("cot_hist_afp", 0)
            elif id_concepto == "sis":
                key_sis = f"{fecha_proceso}{id_institucion}"
                if not cot_afp.empty and "id_afp_hist" in cot_afp.columns:
                    r = cot_afp[cot_afp["id_afp_hist"] == key_sis]
                    if not r.empty:
                        cot_jubilacion = r.iloc[0].get("sis_hist", 0)
            elif id_concepto == "cesEmpleado":
                cot_jubilacion = 0.6
            elif id_concepto == "isapre":
                cot_jubilacion = monto
            elif id_concepto == "mutual":
                if not mutuales.empty and "cod_lre" in mutuales.columns and "nombre_mutual" in mutuales.columns:
                    m = mutuales[mutuales["cod_lre"] == col_1152]
                    if not m.empty:
                        nombre_mutual = m.iloc[0]["nombre_mutual"]
                        if not empresas.empty and "Mutual" in empresas.columns and "Cotización Mutual" in empresas.columns:
                            e = empresas[empresas["Mutual"] == nombre_mutual]
                            if not e.empty:
                                cot_jubilacion = e.iloc[0]["Cotización Mutual"]
            elif id_concepto == "licenciaDias":
                cot_jubilacion = dias_licencia

            filas.append({
                "Fecha de proceso": fecha_proceso,
                "Id empleado": rut,
                "Número de contrato": 1,
                "Id del concepto": id_concepto,
                "Monto del concepto": monto,
                "Afecto": afecto,
                "Id de institución": id_institucion,
                "Cotización de jubilación": cot_jubilacion,
                "Días de licencias": dias_licencia,
                "Días trabajados": dias_vacaciones,
                "Fecha de aplicación": "x",
                "Empresa": empresa_salida,
                "Total de rebajas por L": rebaja_zona,
                "Jornada": "C",
                "Fase": 1,
                "Días de vacaciones": dias_vacaciones,
                "Monto Init": monto_init,
            })

        # Fila adicional licenciaDias si aplica
        if dias_licencia > 0:
            filas.append({
                "Fecha de proceso": fecha_proceso,
                "Id empleado": rut,
                "Número de contrato": 1,
                "Id del concepto": "licenciaDias",
                "Monto del concepto": dias_licencia,
                "Afecto": "",
                "Id de institución": "",
                "Cotización de jubilación": dias_licencia,
                "Días de licencias": dias_licencia,
                "Días trabajados": dias_vacaciones,
                "Fecha de aplicación": "x",
                "Empresa": empresa_salida,
                "Total de rebajas por L": rebaja_zona,
                "Jornada": "C",
                "Fase": 1,
                "Días de vacaciones": dias_vacaciones,
                "Monto Init": monto_init,
            })

    return pd.DataFrame(filas)

def generar_excel(df_salida):
    """Genera el Excel de salida con formato."""
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Migración"

    header_fill = PatternFill("solid", fgColor="1A2744")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(
        bottom=Side(style="thin", color="E8EDF5"),
        right=Side(style="thin", color="E8EDF5")
    )

    cols = list(df_salida.columns)
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = max(len(col) + 4, 14)

    for ri, row in enumerate(df_salida.itertuples(index=False), 2):
        fill = PatternFill("solid", fgColor="EAF0F8") if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    wb.save(output)
    return output.getvalue()

# ─────────────────────────────────────────────
# INTERFAZ PRINCIPAL
# ─────────────────────────────────────────────

# Header
st.markdown("""
<div class="rex-header">
    <div style="display:flex; align-items:center; gap:12px;">
        <div class="rex-logo">Rex<span>+</span></div>
        <span class="rex-title">Migración DDJJ Previred</span>
    </div>
    <div class="rex-badge">PRODUCCIÓN</div>
</div>
""", unsafe_allow_html=True)

# Título sección
st.markdown('<div class="section-title">📂 Migración DDJJ Previred</div>', unsafe_allow_html=True)
st.markdown('<div class="section-sub">Sube uno o más archivos CSV del mismo RUT empresa para generar el archivo de salida en Excel.</div>', unsafe_allow_html=True)

# Cómo funciona
col1, col2, col3 = st.columns(3)
with col1:
    st.markdown("""
    <div class="step-card">
        <div class="step-label">PASO 1</div>
        <div class="step-title">Subir archivos CSV</div>
        <div class="step-desc">Uno o más archivos del mismo RUT empresa. Ej: 76247825-0_202601.csv</div>
    </div>""", unsafe_allow_html=True)
with col2:
    st.markdown("""
    <div class="step-card">
        <div class="step-label">PASO 2</div>
        <div class="step-title">Validación automática</div>
        <div class="step-desc">Se verifican las cuadraturas contables de cada registro.</div>
    </div>""", unsafe_allow_html=True)
with col3:
    st.markdown("""
    <div class="step-card">
        <div class="step-label">PASO 3</div>
        <div class="step-title">Descargar Excel</div>
        <div class="step-desc">Si todo cuadra, se genera el archivo de salida listo para importar.</div>
    </div>""", unsafe_allow_html=True)

st.markdown('<hr class="rex-divider">', unsafe_allow_html=True)

# Cargar referencias
refs, errores_refs = cargar_referencias()
if errores_refs:
    st.markdown(f'<div class="alert-warning">⚠️ Archivos de referencia no encontrados en <b>/data</b>: {", ".join(errores_refs)}</div>', unsafe_allow_html=True)

# Upload
st.markdown("### 📤 Subir archivos CSV")
archivos = st.file_uploader(
    "Selecciona uno o más archivos CSV de Previred",
    type=["csv"],
    accept_multiple_files=True,
    help="Los archivos deben corresponder al mismo RUT empresa (primeros 10 caracteres del nombre)"
)

if archivos:
    st.markdown(f'<div class="alert-success">✅ {len(archivos)} archivo(s) cargado(s): {", ".join([f.name for f in archivos])}</div>', unsafe_allow_html=True)

    # Validar misma empresa
    valido, prefijos = validar_archivos(archivos)
    if not valido:
        st.markdown(f"""
        <div class="alert-error">
            ❌ <b>Los archivos no corresponden a la misma empresa.</b><br>
            Se detectaron distintos RUT empresa: {", ".join(set(prefijos))}<br>
            Por favor sube solo archivos del mismo RUT empresa.
        </div>""", unsafe_allow_html=True)
        st.stop()

    if st.button("▶ Ejecutar validaciones"):
        todos_errores = []
        dfs = []

        with st.spinner("Procesando archivos..."):
            for archivo in archivos:
                for enc in ("utf-8", "latin-1", "utf-8-sig", "cp1252"):
                    try:
                        archivo.seek(0)
                        df = pd.read_csv(archivo, encoding=enc, sep=None, engine="python")
                        break
                    except (UnicodeDecodeError, Exception):
                        continue
                else:
                    st.error(f"❌ No se pudo leer el archivo {archivo.name}. Verifica que sea un CSV válido.")
                    st.stop()
                df = calcular_totales(df)
                errores = validar_cuadraturas(df, archivo.name)
                todos_errores.extend(errores)
                fecha_proceso = extraer_fecha_proceso(archivo.name)
                df["_fecha_proceso"] = fecha_proceso
                dfs.append(df)

        st.markdown('<hr class="rex-divider">', unsafe_allow_html=True)
        st.markdown("### 🔍 Resultado de validaciones")

        if todos_errores:
            st.markdown(f"""
            <div class="alert-error">
                ❌ <b>No se puede generar el archivo de salida.</b><br>
                Se encontraron <b>{len(todos_errores)} error(es)</b> de validación en los registros procesados.
            </div>""", unsafe_allow_html=True)

            with st.expander("📋 Ver log de errores detallado"):
                df_errores = pd.DataFrame(todos_errores)
                st.dataframe(df_errores, use_container_width=True, hide_index=True)

                csv_log = df_errores.to_csv(index=False).encode("utf-8")
                st.download_button(
                    label="⬇️ Descargar log de errores (.csv)",
                    data=csv_log,
                    file_name=f"log_errores_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv"
                )
        else:
            st.markdown("""
            <div class="alert-success">
                ✅ <b>Todas las validaciones se cumplieron correctamente.</b><br>
                Los registros de todos los archivos cuadran sin diferencias.
            </div>""", unsafe_allow_html=True)

            st.markdown("#### ¿Desea generar el archivo de salida?")
            col_a, col_b, col_c, _ = st.columns([1, 1, 1, 4])

            with col_a:
                aceptar = st.button("✅ Aceptar")
            with col_b:
                cancelar = st.button("✖ Cancelar")
            with col_c:
                salir = st.button("🚪 Salir")

            if salir:
                st.markdown('<div class="alert-warning">La sesión ha sido cerrada. Puedes cerrar esta ventana.</div>', unsafe_allow_html=True)
                st.stop()

            if cancelar:
                st.markdown('<div class="alert-warning">Operación cancelada. Puedes subir nuevos archivos.</div>', unsafe_allow_html=True)
                st.stop()

            if aceptar:
                with st.spinner("Generando archivo de salida..."):
                    df_combined = pd.concat(dfs, ignore_index=True)
                    filas_salida = []
                    for _, grupo in df_combined.groupby("_fecha_proceso"):
                        fp = grupo["_fecha_proceso"].iloc[0]
                        df_out = generar_filas_salida(grupo, fp, refs)
                        filas_salida.append(df_out)

                    df_final = pd.concat(filas_salida, ignore_index=True) if filas_salida else pd.DataFrame()
                    excel_bytes = generar_excel(df_final)

                st.markdown('<div class="alert-success">✅ Archivo generado exitosamente.</div>', unsafe_allow_html=True)
                nombre_empresa = archivos[0].name[:10]
                st.download_button(
                    label="⬇️ Descargar archivo de salida (.xlsx)",
                    data=excel_bytes,
                    file_name=f"migracion_{nombre_empresa}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
